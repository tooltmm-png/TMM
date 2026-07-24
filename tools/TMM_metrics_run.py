"""
TMM_metrics_run.py
==================
Pipeline completo do TMM: recebe baseline e versões via CLI,
calcula todas as métricas e gera o xlsx com tabelas e gráficos.

USO
───
python tools/TMM_metrics_run.py \
    --baseline dataset/reports/vulnnet_scans_openvas.csv \
    --versions TMMv3:dataset/extractions/deepseek_v3.csv:deepseek \
               TMMv3:dataset/extractions/gpt4_v3.csv:gpt4 \
    --xlsx     TMM_metrics_v3.xlsx \
    [--db      metrics_db.json]          # opcional: salva o DB para reuso

EXEMPLO COM UMA SÓ VERSÃO
──────────────────────────
python tools/TMM_metrics_run.py \
    --baseline dataset/reports/vulnnet_scans_openvas.csv \
    --versions TMMv3:dataset/extractions/deepseek_v3.csv:deepseek \
    --xlsx     TMM_metrics_v3.xlsx

EXEMPLO COM AS 5 LLMs (V3), BASELINE E DATASETS REAIS DO REPO
───────────────────────────────────────────────────────────────
python tools/TMM_metrics_run.py \
    --baseline baselines/native/vulnnet_scans_openvas.csv \
    --versions TMMv3:artifacts/v3/openvas_129_dockers/deepseek_v3.csv:deepseek \
               TMMv3:artifacts/v3/openvas_129_dockers/gpt4_v3.csv:gpt4 \
               TMMv3:artifacts/v3/openvas_129_dockers/gpt5_v3.csv:gpt5 \
               TMMv3:artifacts/v3/openvas_129_dockers/llama3_v3.csv:llama3 \
               TMMv3:artifacts/v3/openvas_129_dockers/llama4_v3.csv:llama4 \
    --xlsx     artifacts/v3/TMM_metrics_v3.xlsx

DEPENDÊNCIAS
────────────
pip install pandas openpyxl matplotlib numpy rouge-score scikit-learn
"""

import argparse, ast, re, json, sys, io, math, warnings, hashlib, os
from pathlib import Path
from datetime import datetime
from difflib import SequenceMatcher
warnings.filterwarnings('ignore')

# Bump this whenever the merge/computation logic changes to force cache invalidation
COMPUTE_VERSION = '2'

import numpy as np
import pandas as pd
from rouge_score import rouge_scorer as rouge_lib
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.metrics import confusion_matrix, classification_report

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.gridspec import GridSpec
import matplotlib.colors as mcolors
from matplotlib.colors import ListedColormap, BoundaryNorm

matplotlib.rcParams.update({
    'font.family':       'sans-serif',
    'font.sans-serif':   ['Arial', 'DejaVu Sans'],
    'font.size':         10,
    'axes.titlesize':    12,
    'axes.titleweight':  'bold',
    'axes.labelsize':    10,
    'xtick.labelsize':    9,
    'ytick.labelsize':    9,
    'legend.fontsize':   11,
    'legend.framealpha': 0.92,
    'legend.edgecolor':  '#CCCCCC',
    'figure.facecolor':  'white',
    'axes.facecolor':    'white',
    'axes.edgecolor':    '#AAAAAA',
    'axes.linewidth':    0.8,
    'grid.color':        '#E5E5E5',
    'grid.linewidth':    0.5,
    'xtick.major.size':  3,
    'ytick.major.size':  3,
    'xtick.major.width': 0.6,
    'ytick.major.width': 0.6,
    'xtick.direction':   'out',
    'ytick.direction':   'out',
    'lines.linewidth':   1.4,
    'patch.linewidth':   0.5,
})

# Scientific discrete colormap: Divergent→Slightly→Moderately→Highly Similar
_SC_BOUNDS = [0.0, 0.40, 0.60, 0.70, 1.001]
_SC_COLORS = ['#D7191C', '#FDAE61', '#A6D96A', '#1A9641']
_SC_CMAP   = ListedColormap(_SC_COLORS)
_SC_NORM   = BoundaryNorm(_SC_BOUNDS, _SC_CMAP.N)
_SC_ABSENT = '#E0E0E0'
_SC_CAT_LABELS = [
    (0.70, 'Highly Similar (≥0.70)'),
    (0.60, 'Moderately Similar (0.60–0.70)'),
    (0.40, 'Slightly Similar (0.40–0.60)'),
    (0.00, 'Divergent (<0.40)'),
]

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════════════════
# FIELD CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════════

FIELD_MAP = {
    'description':              ('description',              'Summary'),
    'detection_result':         ('detection_result',         'Specific Result'),
    'detection_method':         ('detection_method',         'Vulnerability Detection Method'),
    'product_detection_result': ('product_detection_result', 'Product Detection Result'),
    'impact':                   ('impact',                   'Impact'),
    'solution':                 ('solution',                 'Solution'),
    'insight':                  ('insight',                  'Vulnerability Insight'),
    'references':               ('references',               'CVEs'),
    'cvss':                     ('cvss',                     'CVSS'),
    'port':                     ('port',                     'Port'),
    'protocol':                 ('protocol',                 'Port Protocol'),
    'severity':                 ('severity',                 'Severity'),
    'log_method':               (None, None),
    'plugin_details':           (None, None),
    'instances':                (None, None),
    'source':                   (None, None),
    'plugin':                   (None, None),
}

CAMPO_META = {
    'description':              {'tipo': 'Semântico',      'metrica_principal': 'ROUGE-L'},
    'detection_result':         {'tipo': 'Semântico',      'metrica_principal': 'ROUGE-L'},
    'detection_method':         {'tipo': 'Semântico',      'metrica_principal': 'ROUGE-L'},
    'product_detection_result': {'tipo': 'Semântico',      'metrica_principal': 'ROUGE-L'},
    'impact':                   {'tipo': 'Semântico',      'metrica_principal': 'ROUGE-L'},
    'solution':                 {'tipo': 'Semântico',      'metrica_principal': 'ROUGE-L'},
    'insight':                  {'tipo': 'Semântico',      'metrica_principal': 'ROUGE-L'},
    'references':               {'tipo': 'Semântico',      'metrica_principal': 'Set-F1 (CVE)'},
    'cvss':                     {'tipo': 'Determinístico', 'metrica_principal': 'Exact Match'},
    'port':                     {'tipo': 'Determinístico', 'metrica_principal': 'Exact Match'},
    'protocol':                 {'tipo': 'Determinístico', 'metrica_principal': 'Exact Match'},
    'severity':                 {'tipo': 'Determinístico', 'metrica_principal': 'F1-macro'},
    'log_method':               {'tipo': 'N/A',            'metrica_principal': 'N/A'},
    'plugin_details':           {'tipo': 'N/A',            'metrica_principal': 'N/A'},
    'instances':                {'tipo': 'N/A',            'metrica_principal': 'N/A'},
    'source':                   {'tipo': 'N/A',            'metrica_principal': 'N/A'},
    'plugin':                   {'tipo': 'N/A',            'metrica_principal': 'N/A'},
}

CAMPO_ORDER    = list(FIELD_MAP.keys())
SEVERITY_ORDER = ['LOG', 'LOW', 'MEDIUM', 'HIGH', 'CRITICAL']


# ═══════════════════════════════════════════════════════════════════════════
# STYLE CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════

PALETTE = {
    'dark':  '1F3864', 'white': 'FFFFFF', 'alt':   'F5F7FA',
    'good':  'C8E6C9', 'med':   'FFF9C4', 'bad':   'FFCCBC',
    'blank': 'EEEEEE', 'sem':   '1A237E', 'det':   '1B5E20', 'na': '546E7A',
}

VERSION_COLORS = ['4A148C', '0D47A1', '1B5E20', 'BF360C', '880E4F', '006064']

METRIC_COLORS = {
    'rouge_l':      '#1565C0', 'token_f1':  '#6A1B9A', 'tfidf_cosine': '#00838F',
    'bertscore_f1': '#2E7D32', 'f1_macro':  '#1B5E20',
    'precision':    '#1A237E', 'recall':    '#880E4F',
}

TIPO_EN    = {'Semântico': 'Semantic', 'Determinístico': 'Deterministic', 'N/A': 'N/A'}
METRICA_EN = {'ROUGE-L': 'ROUGE-L', 'Token-F1': 'Token-F1', 'Set-F1 (CVE)': 'Set-F1 (CVE)',
              'Exact Match': 'Exact Match', 'F1-macro': 'F1-macro', 'N/A': 'N/A'}
CAT_LABEL_EN = {
    'Semantic':      '● Semantic Fields',
    'Deterministic': '● Deterministic Fields',
    'N/A':           '● Fields Without Direct CSV Column',
}
OBS_EN = {
    'cvss':     'Direct numeric comparison',
    'port':     'Direct numeric comparison',
    'protocol': 'tcp/udp — case insensitive',
    'severity': 'LOG/LOW/MEDIUM/HIGH/CRITICAL — see Confusion Matrix tab',
}

BORDER_THIN = Border(
    left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin',  color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF'),
)


# ═══════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════

def vc(i):  return VERSION_COLORS[i % len(VERSION_COLORS)]
def vch(i): return '#' + vc(i)

# Score categories: (min_threshold, hex_color, label)
# Highly Similar ≥ 0.70 | Moderately Similar 0.60–0.70 |
# Slightly Similar 0.40–0.60 | Divergent < 0.40 | Absent = N/A
SCORE_CATS = [
    (0.70, '#C8E6C9', 'Highly Similar'),
    (0.60, '#DCEDC8', 'Moderately Similar'),
    (0.40, '#FFF9C4', 'Slightly Similar'),
    (0.00, '#FFCCBC', 'Divergent'),
]
SCORE_ABSENT = '#EEEEEE'

def score_bg(v):
    if v is None: return SCORE_ABSENT
    for thr, color, _ in SCORE_CATS:
        if v >= thr:
            return color
    return SCORE_CATS[-1][1]

def score_bg_clean(v):
    return score_bg(v).replace('#', '')

def bc(ws, row, col, val=None, bold=False, italic=False, halign='center',
       bg=None, fg='000000', size=12, wrap=True, border=True):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=bold, italic=italic, color=fg, name='Arial', size=size)
    c.alignment = Alignment(horizontal=halign, vertical='center', wrap_text=wrap)
    if border: c.border = BORDER_THIN
    if bg:     c.fill = PatternFill('solid', start_color=bg)
    return c

# ── Vector figure export (paper) ───────────────────────────────────────────
# Set via --figdir/--paper-layout: every chart is also written as PDF+SVG
# (vector, tight bbox) under FIGDIR, named by the chart slug.
FIGDIR = None
PAPER_LAYOUT = False
_CURRENT_FIG_NAME = None

def export_vector(fig, name=None):
    """Save fig as <FIGDIR>/<name>.pdf and .svg (vector). No-op without --figdir."""
    name = name or _CURRENT_FIG_NAME
    if not (FIGDIR and name):
        return
    for ext in ('.pdf', '.svg'):
        fig.savefig(Path(FIGDIR) / f'{name}{ext}',
                    bbox_inches='tight', facecolor='white')

def fig_to_buf(fig, dpi=180):
    export_vector(fig)
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=dpi, bbox_inches='tight', facecolor='white')
    buf.seek(0)
    return buf

def parse_list(val):
    if pd.isna(val) or str(val).strip() in ['[]', 'nan', '']: return []
    try:    return [str(x) for x in ast.literal_eval(str(val)) if x]
    except: return [str(val)]

def list_to_text(val): return ' '.join(parse_list(val))

def clean(val):
    if pd.isna(val) or str(val).strip() in ['nan', '']: return ''
    return re.sub(r'\s+', ' ', str(val).replace('\r\n', ' ').replace('\n', ' ')).strip()

def extract_cves(text):
    return set(re.findall(r'CVE-\d{4}-\d+', str(text), re.IGNORECASE))

def stats(arr):
    return {'mean': round(float(arr.mean()), 4),
            'min':  round(float(arr.min()),  4),
            'max':  round(float(arr.max()),  4)}

def principal_score(campo, r):
    if not r or r.get('ausente'): return None
    tipo = CAMPO_META.get(campo, {}).get('tipo')
    if campo == 'references':    return r.get('set_f1_cve', {}).get('mean')
    if campo == 'severity':      return r.get('f1_macro')
    if tipo == 'Semântico':      return r.get('rouge_l', {}).get('mean')
    if tipo == 'Determinístico': return r.get('f1') or r.get('exact_match_rate')
    return None


# ═══════════════════════════════════════════════════════════════════════════
# METRICS COMPUTATION
# ═══════════════════════════════════════════════════════════════════════════

def soft_f1(hyps, refs):
    corpus = hyps + refs
    vect = TfidfVectorizer(analyzer='word', min_df=1, token_pattern=r'(?u)\b\w+\b')
    vect.fit(corpus)
    P, R, F = [], [], []
    for h, r in zip(hyps, refs):
        hw = (h or '').lower().split() or ['']
        rw = (r or '').lower().split() or ['']
        sim = cosine_similarity(vect.transform(hw), vect.transform(rw))
        p  = float(sim.max(axis=1).mean())
        rv = float(sim.max(axis=0).mean())
        f  = 2*p*rv/(p+rv) if (p+rv) > 0 else 0.0
        P.append(p); R.append(rv); F.append(f)
    return np.array(P), np.array(R), np.array(F)


_NAME_MAP_CACHE: dict = {}  # frozenset(llm_names) -> {llm_name: bl_name}

def _build_name_map(llm_names, bl_names, threshold=0.85):
    """Map each LLM Name to the most similar baseline NVT Name (>= threshold).
    Fast path: exact match, then quote-stripped exact match.
    Slow path (SequenceMatcher) only for the few remaining unmatched names.
    Results are memoised per unique set of llm_names."""
    cache_key = frozenset(llm_names)
    if cache_key in _NAME_MAP_CACHE:
        return _NAME_MAP_CACHE[cache_key]

    bl_set = set(bl_names)
    # stripped baseline index: remove surrounding/embedded quotes for fast lookup
    def _strip_quotes(s):
        return s.replace("'", "").replace('"', "").strip()
    bl_stripped = {_strip_quotes(b): b for b in bl_names}

    name_map = {}
    unmatched = []
    for ln in llm_names:
        if ln in bl_set:
            name_map[ln] = ln
        else:
            sq = _strip_quotes(ln)
            if sq in bl_set:
                name_map[ln] = sq
            elif sq in bl_stripped:
                name_map[ln] = bl_stripped[sq]
            else:
                unmatched.append(ln)

    # Only run expensive SequenceMatcher on truly unmatched names
    if unmatched:
        bl_list = list(bl_names)
        for ln in unmatched:
            best_score, best_name = 0.0, None
            for bn in bl_list:
                s = SequenceMatcher(None, ln.lower(), bn.lower()).ratio()
                if s > best_score:
                    best_score, best_name = s, bn
                    if s == 1.0:
                        break
            if best_score >= threshold:
                name_map[ln] = best_name

    _NAME_MAP_CACHE[cache_key] = name_map
    return name_map


def compute_version(merged_path, baseline_path, version_label):
    print(f'\n  [{version_label}] Loading data...')
    merged   = pd.read_csv(merged_path,   encoding='utf-8-sig')
    baseline = pd.read_csv(baseline_path, encoding='utf-8-sig')
    merged.columns   = [c.strip().lstrip('﻿') for c in merged.columns]
    baseline.columns = [c.strip().lstrip('﻿') for c in baseline.columns]

    # ── Build name mapping: LLM Name → baseline NVT Name (≥85% similarity) ──
    mg_names = merged['Name'].dropna().astype(str).str.strip().unique()
    bl_names = baseline['NVT Name'].dropna().astype(str).str.strip().unique()
    name_map = _build_name_map(mg_names, bl_names, threshold=0.85)
    print(f'  [{version_label}] Name mapping: {len(name_map)}/{len(mg_names)} LLM names matched')

    # Apply mapping and normalise host keys
    merged   = merged.copy()
    baseline = baseline.copy()
    merged['_mapped_name'] = merged['Name'].astype(str).str.strip().map(name_map)
    merged['_host_key']    = merged['host'].astype(str).str.strip()
    baseline['_name_key']  = baseline['NVT Name'].astype(str).str.strip()
    baseline['_ip_key']    = baseline['IP'].astype(str).str.strip()

    # Deduplicate: one row per (host, mapped_name) and per (IP, NVT Name)
    merged_agg   = (merged.dropna(subset=['_mapped_name'])
                          .sort_values(['_host_key', '_mapped_name'])
                          .groupby(['_host_key', '_mapped_name'], sort=False)
                          .first().reset_index())
    baseline_agg = (baseline.sort_values(['_ip_key', '_name_key'])
                            .groupby(['_ip_key', '_name_key'], sort=False)
                            .first().reset_index())

    # Drop overlapping columns to avoid _x/_y suffixes
    overlap = set(merged_agg.columns) & set(baseline_agg.columns) - \
              {'_host_key', '_mapped_name', '_ip_key', '_name_key'}
    if overlap:
        merged_agg = merged_agg.drop(columns=list(overlap), errors='ignore')

    df = merged_agg.merge(baseline_agg,
                          left_on=['_host_key', '_mapped_name'],
                          right_on=['_ip_key',  '_name_key'],
                          how='inner')
    N  = len(df)
    print(f'  [{version_label}] Pairs: {N} (host+name matched)')

    rl_scorer = rouge_lib.RougeScorer(['rougeL'], use_stemmer=False)
    results   = {}

    # ── Semantic fields ──────────────────────────────────────────────────
    sem_campos = [c for c in CAMPO_ORDER
                  if CAMPO_META.get(c,{}).get('tipo') == 'Semântico' and c != 'references']

    for campo in sem_campos:
        col_m, col_b = FIELD_MAP[campo]
        if col_m not in df.columns or col_b not in df.columns:
            results[campo] = {'tipo': 'Semântico', 'n_pairs': 0, 'ausente': True}
            continue

        hyps = [list_to_text(r[col_m]) or ' ' for _, r in df.iterrows()]
        refs = [clean(r[col_b])         or ' ' for _, r in df.iterrows()]

        rl_arr  = np.array([rl_scorer.score(ref, hyp)['rougeL'].fmeasure
                            for hyp, ref in zip(hyps, refs)])
        tf1_arr = np.array([_token_f1(h, r) for h, r in zip(hyps, refs)])
        cos_arr = _tfidf_cosine(hyps, refs)
        P, R, F = soft_f1(hyps, refs)

        results[campo] = {
            'tipo': 'Semântico', 'n_pairs': N,
            'rouge_l':        stats(rl_arr),
            'token_f1':       stats(tf1_arr),
            'tfidf_cosine':   stats(cos_arr),
            'bertscore_proxy': {
                'method':    'Soft-F1 (TF-IDF word vectors)',
                'precision': round(float(P.mean()), 4),
                'recall':    round(float(R.mean()), 4),
                'f1_mean':   round(float(F.mean()), 4),
                'f1_min':    round(float(F.min()),  4),
                'f1_max':    round(float(F.max()),  4),
            }
        }
        print(f'  [{version_label}] {campo:30s} ROUGE-L={results[campo]["rouge_l"]["mean"]:.4f}  '
              f'Soft-F1={results[campo]["bertscore_proxy"]["f1_mean"]:.4f}')

    # ── References Set-F1 ────────────────────────────────────────────────
    col_m, col_b = FIELD_MAP['references']
    if col_m in df.columns and col_b in df.columns:
        sf1_vals = []
        for _, r in df.iterrows():
            pred = extract_cves(list_to_text(r[col_m]))
            gold = extract_cves(clean(r[col_b]))
            if not pred and not gold:   sf1_vals.append(1.0)
            elif not pred or not gold:  sf1_vals.append(0.0)
            else:
                inter = pred & gold
                p = len(inter)/len(pred); rc = len(inter)/len(gold)
                sf1_vals.append(2*p*rc/(p+rc) if (p+rc) else 0.0)
        results['references'] = {
            'tipo': 'Semântico', 'n_pairs': N,
            'set_f1_cve': stats(np.array(sf1_vals)),
        }
        print(f'  [{version_label}] references/Set-F1 CVE: mean={results["references"]["set_f1_cve"]["mean"]:.4f}')

    # ── Deterministic: cvss ──────────────────────────────────────────────
    col_m, col_b = FIELD_MAP['cvss']
    if col_m in df.columns and col_b in df.columns:
        # V2 stores cvss as list-string e.g. "['9.8']" — unwrap first element
        def _unwrap_cvss(val):
            items = parse_list(val)
            return items[0] if items else None
        vm = pd.to_numeric(df[col_m].apply(_unwrap_cvss), errors='coerce')
        vb = pd.to_numeric(df[col_b], errors='coerce')
        mask = vm.notna() & vb.notna()
        em = float((vm[mask] == vb[mask]).mean())
        results['cvss'] = {'tipo':'Determinístico','n_pairs':int(mask.sum()),
                           'exact_match_rate':round(em,4),'precision':round(em,4),
                           'recall':round(em,4),'f1':round(em,4)}
        print(f'  [{version_label}] cvss Exact Match: {em:.4f}')

    # ── Deterministic: port ──────────────────────────────────────────────
    col_m, col_b = FIELD_MAP['port']
    if col_m in df.columns and col_b in df.columns:
        # LLMs store port as float (80.0); convert both sides to int before comparing
        pm = pd.to_numeric(df[col_m], errors='coerce').fillna(-1).astype(int).astype(str)
        pb = df[col_b].fillna(0).astype(int).astype(str)
        mask = (pm != '-1') & (pb != '0')
        em = float((pm[mask].values == pb[mask].values).mean()) if mask.sum() > 0 else 0.0
        results['port'] = {'tipo':'Determinístico','n_pairs':int(mask.sum()),
                           'exact_match_rate':round(em,4),'precision':round(em,4),
                           'recall':round(em,4),'f1':round(em,4)}
        print(f'  [{version_label}] port Exact Match: {em:.4f}')

    # ── Deterministic: protocol ──────────────────────────────────────────
    col_m, col_b = FIELD_MAP['protocol']
    if col_m in df.columns and col_b in df.columns:
        pm = df[col_m].str.lower().str.strip()
        pb = df[col_b].str.lower().str.strip()
        mask = pm.notna() & pb.notna()
        em = float((pm[mask].values == pb[mask].values).mean())
        results['protocol'] = {'tipo':'Determinístico','n_pairs':int(mask.sum()),
                               'exact_match_rate':round(em,4),'precision':round(em,4),
                               'recall':round(em,4),'f1':round(em,4)}
        print(f'  [{version_label}] protocol Exact Match: {em:.4f}')

    # ── Deterministic: severity ──────────────────────────────────────────
    col_m, col_b = FIELD_MAP['severity']
    if col_m in df.columns and col_b in df.columns:
        sm = df[col_m].str.upper().str.strip()
        sb = df[col_b].str.upper().str.strip()
        mask = sm.notna() & sb.notna()
        sm_v, sb_v = sm[mask].values, sb[mask].values
        em = float((sm_v == sb_v).mean())
        labels = [s for s in SEVERITY_ORDER if s in set(sm_v) | set(sb_v)]
        cr = classification_report(sb_v, sm_v, labels=labels, output_dict=True, zero_division=0)
        cm_matrix = confusion_matrix(sb_v, sm_v, labels=labels)
        results['severity'] = {
            'tipo': 'Determinístico', 'n_pairs': int(mask.sum()),
            'exact_match_rate':  round(em, 4),
            'precision_macro':   round(cr['macro avg']['precision'], 4),
            'recall_macro':      round(cr['macro avg']['recall'],    4),
            'f1_macro':          round(cr['macro avg']['f1-score'],  4),
            'confusion_matrix': {
                'labels': labels, 'matrix': cm_matrix.tolist(),
                'per_class': {k: {m: round(v, 4) for m, v in vv.items()}
                              for k, vv in cr.items()
                              if k not in ['accuracy', 'macro avg', 'weighted avg']},
            }
        }
        print(f'  [{version_label}] severity Exact={em:.4f}  F1-macro={results["severity"]["f1_macro"]:.4f}')

    # ── N/A fields ───────────────────────────────────────────────────────
    for campo in ['log_method', 'plugin_details', 'instances', 'source', 'plugin']:
        results[campo] = {'tipo': 'N/A', 'n_pairs': 0}

    return results


def _token_f1(h, r):
    h_t = set(h.lower().split()); r_t = set(r.lower().split())
    common = h_t & r_t
    if not common: return 0.0
    p = len(common)/len(h_t); rc = len(common)/len(r_t)
    return 2*p*rc/(p+rc)

def _tfidf_cosine(hyps, refs):
    vect = TfidfVectorizer(min_df=1)
    vect.fit(hyps + refs)
    H = vect.transform(hyps); R = vect.transform(refs)
    return np.array([cosine_similarity(H[i], R[i])[0][0] for i in range(len(hyps))])


# ═══════════════════════════════════════════════════════════════════════════
# CHART HELPERS
# ═══════════════════════════════════════════════════════════════════════════

def discrete_heatmap(ax, matrix, xlabels, ylabels, cell=1.0,
                     color_fn=None, val_fmt=None, fontsize=9.0):
    n_row, n_col = len(ylabels), len(xlabels)
    for i in range(n_row):
        for j in range(n_col):
            val   = matrix[i][j]
            color = color_fn(i, j, val) if color_fn else score_bg(val)
            x, y  = j * cell, (n_row - 1 - i) * cell
            ax.add_patch(plt.Rectangle((x, y), cell, cell,
                                       facecolor=color, edgecolor='white',
                                       linewidth=0.8, zorder=2))
            if val_fmt is not None:
                ax.text(x + cell/2, y + cell/2, val_fmt(val),
                        ha='center', va='center', fontsize=fontsize,
                        fontweight='bold', color='#333333', zorder=3)
    ax.set_xlim(0, n_col * cell); ax.set_ylim(0, n_row * cell)
    ax.set_xticks([j + 0.5 for j in range(n_col)])
    ax.set_xticklabels(xlabels, rotation=35, ha='right', fontsize=10.0)
    ax.set_yticks([i + 0.5 for i in range(n_row)])
    ax.set_yticklabels(list(reversed(ylabels)), fontsize=11.0, fontweight='bold')
    ax.tick_params(length=0, pad=3); ax.spines[:].set_visible(False); ax.set_aspect('equal')

def sci_heatmap(ax, fig, data_2d, col_labels, row_labels, fmt='.2f',
                ann_fontsize=8, add_colorbar=True,
                label_fontsize=None, cbar_fontsize=7.5):
    """
    Publication-quality heatmap using imshow.
    data_2d: list[list] or 2D array, None = absent cell.
    Returns the imshow AxesImage.
    """
    n_row, n_col = len(row_labels), len(col_labels)
    mat = np.full((n_row, n_col), np.nan)
    for i in range(n_row):
        for j in range(n_col):
            v = data_2d[i][j]
            if v is not None and not (isinstance(v, float) and np.isnan(v)):
                mat[i, j] = float(v)

    display = np.where(np.isnan(mat), 0.0, mat)
    im = ax.imshow(display, cmap=_SC_CMAP, norm=_SC_NORM,
                   aspect='auto', interpolation='nearest')

    # Absent cells overlay
    for i in range(n_row):
        for j in range(n_col):
            if np.isnan(mat[i, j]):
                ax.add_patch(plt.Rectangle((j-0.5, i-0.5), 1, 1,
                             facecolor=_SC_ABSENT, edgecolor='white', linewidth=0.5, zorder=2))
                ax.text(j, i, 'N/A', ha='center', va='center',
                        fontsize=ann_fontsize-1, color='#888888', zorder=3)
            else:
                val = mat[i, j]
                rgba = _SC_CMAP(_SC_NORM(val))
                bright = 0.299*rgba[0] + 0.587*rgba[1] + 0.114*rgba[2]
                tc = 'white' if bright < 0.50 else '#1A1A1A'
                ax.text(j, i, format(val, fmt), ha='center', va='center',
                        fontsize=ann_fontsize, fontweight='bold', color=tc, zorder=3)

    # White cell borders
    for x in np.arange(-0.5, n_col, 1):
        ax.axvline(x, color='white', linewidth=0.6, zorder=4)
    for y in np.arange(-0.5, n_row, 1):
        ax.axhline(y, color='white', linewidth=0.6, zorder=4)

    ax.set_xticks(range(n_col))
    ax.set_xticklabels(col_labels, rotation=40, ha='right',
                       **({'fontsize': label_fontsize} if label_fontsize else {}))
    ax.set_yticks(range(n_row))
    ax.set_yticklabels(row_labels,
                       **({'fontsize': label_fontsize} if label_fontsize else {}))
    ax.tick_params(length=0, pad=3)
    ax.spines[:].set_visible(False)

    if add_colorbar:
        cbar = fig.colorbar(im, ax=ax, fraction=0.03, pad=0.02, aspect=25,
                            boundaries=_SC_BOUNDS, ticks=[0.20, 0.50, 0.65, 0.855])
        cbar.ax.set_yticklabels(
            ['Divergent\n(<0.40)', 'Slightly\nSimilar\n(0.40-0.60)',
             'Moderately\nSimilar\n(0.60-0.70)', 'Highly\nSimilar\n(>=0.70)'],
            fontsize=cbar_fontsize)
        cbar.ax.tick_params(length=0, pad=4)
        cbar.outline.set_visible(False)
    return im


def make_chart_sheet(wb, title, subtitle, bg, buf, img_w=800, img_h=333):
    ws = wb.create_sheet(title)
    ws.merge_cells('A1:P1')
    c = ws['A1']; c.value = title.replace('📊 ','').replace('📈 ','')
    c.font = Font(bold=True, color='FFFFFF', name='Arial', size=16)
    c.fill = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 34
    ws.merge_cells('A2:P2')
    c2 = ws['A2']; c2.value = subtitle
    c2.font = Font(italic=True, color='444444', name='Arial', size=12)
    c2.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18
    img = XLImage(buf); img.width = img_w; img.height = img_h
    ws.add_image(img, 'A4')
    for r in range(4, 60): ws.row_dimensions[r].height = 15
    ws.column_dimensions['A'].width = 15
    ws.sheet_view.showGridLines = False
    return ws


# ═══════════════════════════════════════════════════════════════════════════
# CHART GENERATORS
# ═══════════════════════════════════════════════════════════════════════════


def _ver_llm_label(ver, version_llms, multiline=False):
    """Single label with LLM — used in titles and legends."""
    llm = (version_llms or {}).get(ver)
    if not llm:
        return ver
    return f'{llm}\n{ver}' if multiline else f'{llm} {ver}'


def _build_ver_labels(versions, version_llms, multiline=True):
    """
    Tick labels: LLM shown once per group, centred on the middle version.
    e.g. 3 versions of deepseek → ['V1', 'deepseek\nV2', 'V3']
    """
    llms   = [(version_llms or {}).get(v) for v in versions]
    labels = list(versions)
    i = 0
    while i < len(versions):
        llm = llms[i]
        if not llm:
            i += 1; continue
        j = i
        while j < len(versions) and llms[j] == llm:
            j += 1
        mid = i + (j - i - 1) // 2
        for k in range(i, j):
            labels[k] = (f'{llm}\n{versions[k]}' if multiline else f'{llm} {versions[k]}') if k == mid else versions[k]
        i = j
    return labels

_VER_LABEL_COLORS = {'TMMv1': '#E57373', 'TMMv2': '#FFC107', 'TMMv3': '#4DD0E1'}


def _draw_llm_groups(ax, versions, version_llms):
    """Two-tier y-axis labels for heatmaps grouped by LLM.

    Tier 1 (close to heatmap)  – version label (V1/V2/V3) in distinct colors.
    Tier 2 (further left)      – LLM name horizontal, bold, centred on its rows.
    """
    n_ver = len(versions)
    ax.set_yticklabels([''] * n_ver)
    ax.tick_params(axis='y', length=0, pad=1)

    # Build LLM groups (consecutive rows sharing the same LLM)
    cur_llm, llm_start = None, 0
    groups = []
    for i, ver in enumerate(versions):
        llm = version_llms.get(ver, ver.split('|')[0] if '|' in ver else ver)
        if llm != cur_llm:
            if cur_llm is not None:
                groups.append((llm_start, i - 1, cur_llm))
                ax.axhline(i - 0.5, color='#555555', linewidth=1.6, zorder=6)
            llm_start, cur_llm = i, llm
    if cur_llm is not None:
        groups.append((llm_start, n_ver - 1, cur_llm))

    _ver_fs = 13 if PAPER_LAYOUT else 9.0
    _llm_fs = 14 if PAPER_LAYOUT else 10
    # Tier 1: version labels with distinct colors, close to the heatmap
    for i, ver in enumerate(versions):
        ver_label = ver.split('|')[-1] if '|' in ver else ver
        col = _VER_LABEL_COLORS.get(ver_label, '#555555')
        y_frac = 1.0 - (i + 0.5) / n_ver
        ax.text(-0.018, y_frac, ver_label,
                ha='right', va='center', fontsize=_ver_fs, color=col,
                fontweight='bold', transform=ax.transAxes, clip_on=False)

    # Tier 2: LLM name horizontal (no rotation), bold, further left.
    # Afastado (~-0.30) para não colar no rótulo de versão (tier 1) com fonte grande.
    _llm_x = -0.30 if PAPER_LAYOUT else -0.12
    for start, end, llm in groups:
        mid_row = (start + end) / 2
        y_frac = 1.0 - (mid_row + 0.5) / n_ver
        ax.text(_llm_x, y_frac, llm,
                ha='right', va='center',
                fontsize=_llm_fs, fontweight='bold', color='#1A1A1A',
                transform=ax.transAxes, clip_on=False)


def chart_overview(versions_data, versions, all_campos, version_llms=None):
    n_ver = len(versions); n_campo = len(all_campos)
    version_llms = version_llms or {}
    # Paper: células largas o suficiente para os números "0.89" respirarem
    # (sem colar nas bordas), e proporção ~quadrada por célula (não esticada).
    cell_w = max(0.90 if PAPER_LAYOUT else 0.48, 8.0 / max(n_campo, 1))
    cell_h = max(0.56 if PAPER_LAYOUT else 0.40, 6.0 / max(n_ver, 1))
    fig_w  = max(n_campo * cell_w + (4.6 if PAPER_LAYOUT else 4.0), 13)
    fig_h  = max(n_ver   * cell_h + (1.6 if PAPER_LAYOUT else 2.0),  5.0)

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    fig.subplots_adjust(left=0.32, right=0.75, top=0.93, bottom=0.18)

    data = [[principal_score(c, versions_data[v].get(c, {}))
             for c in all_campos] for v in versions]
    y_labels = [ver.split('|')[-1] if '|' in ver else ver for ver in versions]

    if PAPER_LAYOUT:
        sci_heatmap(ax, fig, data, all_campos, y_labels, fmt='.2f',
                    ann_fontsize=12, add_colorbar=True,
                    label_fontsize=13, cbar_fontsize=11)
    else:
        sci_heatmap(ax, fig, data, all_campos, y_labels,
                    fmt='.2f', ann_fontsize=max(5.5, min(7.5, 60/max(n_campo, n_ver))),
                    add_colorbar=True)

    _draw_llm_groups(ax, versions, version_llms)
    buf = fig_to_buf(fig, dpi=200); plt.close('all')
    return buf, 950, round(950 * fig_h / fig_w)


def chart_overview_version(ver_label, versions_data, versions, all_campos, version_llms=None):
    ver_subset = [v for v in versions if v.split('|')[-1] == ver_label]
    n_ver = len(ver_subset); n_campo = len(all_campos)
    version_llms = version_llms or {}
    cell_w = max(0.48, 8.0 / max(n_campo, 1))
    cell_h = max(0.55, 6.0 / max(n_ver, 1))
    fig_w  = max(n_campo * cell_w + 4.0, 13)
    fig_h  = max(n_ver   * cell_h + 2.0,  5.0)

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    fig.subplots_adjust(left=0.18, right=0.74, top=0.91, bottom=0.22)

    data = [[principal_score(c, versions_data[v].get(c, {}))
             for c in all_campos] for v in ver_subset]
    y_labels = [version_llms.get(v, v.split('|')[0]) for v in ver_subset]

    sci_heatmap(ax, fig, data, all_campos, y_labels,
                fmt='.2f', ann_fontsize=max(6.0, min(8.0, 60 / max(n_campo, n_ver))),
                add_colorbar=True)
    for tick in ax.get_yticklabels():
        tick.set_fontweight('bold')
        tick.set_fontsize(10)
        tick.set_color('#1A1A1A')
    ax.tick_params(axis='y', length=0, pad=6)
    buf = fig_to_buf(fig, dpi=200); plt.close('all')
    return buf, 950, round(950 * fig_h / fig_w)


def chart_rouge_l(versions_data, versions, sem_campos, version_llms=None):
    import matplotlib.patches as mpatches
    n_ver = len(versions); n_sem = len(sem_campos)
    bw    = min(0.7 / max(n_ver, 1), 0.22)
    x     = np.arange(n_sem)
    LLM_PAL = ['#1565C0','#2E7D32','#6A1B9A','#D84315','#00695C',
               '#4E342E','#37474F','#1A237E','#880E4F','#E65100',
               '#004D40','#BF360C','#0D47A1','#1B5E20','#880E4F']

    # Group versions by LLM for coloring
    llm_list, seen_llms = [], set()
    for v in versions:
        llm = (version_llms or {}).get(v, v.split('|')[0] if '|' in v else v)
        if llm not in seen_llms: llm_list.append(llm); seen_llms.add(llm)
    llm_color = {l: LLM_PAL[i % len(LLM_PAL)] for i, l in enumerate(llm_list)}

    fig_h = max(n_ver * 0.38 + 5.0, 7.0)
    fig   = plt.figure(figsize=(20, fig_h))
    bot   = 0.20; top = 0.88
    ax1   = fig.add_axes([0.04, bot, 0.38, top - bot])
    ax2   = fig.add_axes([0.62, bot, 0.32, top - bot])

    for vi, ver in enumerate(versions):
        vd   = versions_data.get(ver, {})
        llm  = (version_llms or {}).get(ver, ver.split('|')[0] if '|' in ver else ver)
        lbl  = _ver_llm_label(ver, version_llms)
        col  = llm_color.get(llm, LLM_PAL[vi % len(LLM_PAL)])
        off  = (vi - n_ver/2 + 0.5) * bw
        means = [vd.get(c, {}).get('rouge_l', {}).get('mean', 0) or 0 for c in sem_campos]
        mins  = [vd.get(c, {}).get('rouge_l', {}).get('min',  0) or 0 for c in sem_campos]
        maxs  = [vd.get(c, {}).get('rouge_l', {}).get('max',  0) or 0 for c in sem_campos]
        ax1.bar(x + off, means, bw * 0.92, label=lbl, color=col, alpha=0.85,
                yerr=[[m-mn for m,mn in zip(means,mins)],
                      [mx-m  for m,mx in zip(means,maxs)]],
                capsize=3, error_kw={'elinewidth': 0.8, 'ecolor': '#444444', 'alpha': 0.6})
    import matplotlib.lines as mlines, matplotlib.patches as mpatches
    ax1.axhline(0.70, color='#1A9641', linestyle='--', linewidth=0.9, alpha=0.75)
    ax1.axhline(0.60, color='#A6D96A', linestyle='--', linewidth=0.9, alpha=0.75)
    ax1.axhline(0.40, color='#FDAE61', linestyle='--', linewidth=0.9, alpha=0.75)
    ax1.set_xticks(x)
    ax1.set_xticklabels(sem_campos, rotation=35, ha='right')
    ax1.set_ylim(0, 1.08); ax1.set_ylabel('ROUGE-L (mean ± range)')
    ver_handles = ax1.get_legend_handles_labels()[0]
    cat_handles = [
        mlines.Line2D([], [], color='#1A9641', linestyle='--', linewidth=1.2, label='Highly Similar (≥0.70)'),
        mlines.Line2D([], [], color='#A6D96A', linestyle='--', linewidth=1.2, label='Moderately Similar (0.60–0.70)'),
        mlines.Line2D([], [], color='#FDAE61', linestyle='--', linewidth=1.2, label='Slightly Similar (0.40–0.60)'),
        mpatches.Patch(color='#D7191C', alpha=0.55, label='Divergent (<0.40)'),
    ]
    ax1.legend(handles=ver_handles + cat_handles, fontsize=11,
               loc='upper left', bbox_to_anchor=(1.02, 1.0),
               borderaxespad=0, ncol=1, framealpha=0.9)
    ax1.grid(axis='y', alpha=0.4, linewidth=0.5)
    ax1.spines[['top', 'right']].set_visible(False)

    # Heatmap
    data     = [[versions_data.get(v, {}).get(c, {}).get('rouge_l', {}).get('mean')
                 for c in sem_campos] for v in versions]
    y_labels = [ver.split('|')[-1] if '|' in ver else ver for ver in versions]
    sci_heatmap(ax2, fig, data, sem_campos, y_labels, fmt='.2f', ann_fontsize=7.5, add_colorbar=True)
    _draw_llm_groups(ax2, versions, version_llms or {})

    buf = fig_to_buf(fig, dpi=180); plt.close('all')
    return buf, 900, round(900 * fig_h / 20)


def chart_token_f1(versions_data, versions, sem_campos):
    n_ver = len(versions); x = np.arange(len(sem_campos)); bw = 0.8/max(n_ver,1)
    fig, axes = plt.subplots(1, 2, figsize=(18, 7), facecolor='#FAFAFA')
    fig.subplots_adjust(wspace=0.45, left=0.06, right=0.72, top=0.88, bottom=0.16)
    ax1 = axes[0]; ax1.set_facecolor('#F5F7FA')
    for vi, ver in enumerate(versions):
        vd = versions_data[ver]
        means = [vd.get(c,{}).get('token_f1',{}).get('mean',0) or 0 for c in sem_campos]
        mins  = [vd.get(c,{}).get('token_f1',{}).get('min',0)  or 0 for c in sem_campos]
        maxs  = [vd.get(c,{}).get('token_f1',{}).get('max',0)  or 0 for c in sem_campos]
        off   = (vi-(n_ver-1)/2)*bw
        ax1.bar(x+off, means, bw*0.9, label=ver, color=vch(vi), alpha=0.85,
                yerr=[[m-mn for m,mn in zip(means,mins)],[mx-m for m,mx in zip(means,maxs)]],
                capsize=3, error_kw={'elinewidth':1.2,'ecolor':'#555555','alpha':0.7})
    ax1.axhline(0.70, color='#2E7D32', linestyle='--', linewidth=1.0, alpha=0.6, label='Highly Similar (0.70)')
    ax1.axhline(0.60, color='#8BC34A', linestyle='--', linewidth=1.0, alpha=0.6, label='Moderately Similar (0.60)')
    ax1.axhline(0.40, color='#F9A825', linestyle='--', linewidth=1.0, alpha=0.6, label='Slightly Similar (0.40)')
    ax1.set_xticks(x); ax1.set_xticklabels(sem_campos, rotation=30, ha='right', fontsize=11.0)
    ax1.set_ylim(0, 1.12); ax1.set_ylabel('Token-F1 (mean ± min/max)', fontsize=12.0)
    ax1.grid(axis='y', alpha=0.4); ax1.spines[['top','right']].set_visible(False)
    ax2 = axes[1]; ax2.remove(); ax2 = fig.add_subplot(1, 2, 2, projection='polar')
    N = len(sem_campos); angles = [n/float(N)*2*math.pi for n in range(N)]; angles += angles[:1]
    ax2.set_facecolor('#F5F7FA'); ax2.set_theta_offset(math.pi/2); ax2.set_theta_direction(-1)
    ax2.set_xticks(angles[:-1]); ax2.set_xticklabels(sem_campos, size=10)
    ax2.set_ylim(0, 1); ax2.set_yticks([0.2, 0.4, 0.6, 0.8, 1.0])
    ax2.set_yticklabels(['0.2','0.4','0.6','0.8','1.0'], size=9)
    for vi, ver in enumerate(versions):
        vd   = versions_data[ver]
        vals = [vd.get(c,{}).get('token_f1',{}).get('mean',0) or 0 for c in sem_campos]
        vals += vals[:1]
        ax2.plot(angles, vals, linewidth=2, color=vch(vi), label=ver)
        ax2.fill(angles, vals, alpha=0.15, color=vch(vi))
    ax2.legend(loc='upper left', bbox_to_anchor=(1.25, 1.15), fontsize=11.0,
               framealpha=0.9, borderaxespad=0)
    buf = fig_to_buf(fig); plt.close('all')
    return buf, 800, 312


def chart_soft_f1(versions_data, versions, sem_campos):
    n_ver = len(versions); x = np.arange(len(sem_campos)); bw = 0.8/max(n_ver,1); lv = versions[-1]
    fig, axes = plt.subplots(1, 2, figsize=(18, 7), facecolor='#FAFAFA')
    fig.subplots_adjust(wspace=0.45, left=0.06, right=0.72, top=0.88, bottom=0.16)
    ax1 = axes[0]; ax1.set_facecolor('#F5F7FA')
    for vi, ver in enumerate(versions):
        vd  = versions_data[ver]
        f1s = [vd.get(c,{}).get('bertscore_proxy',{}).get('f1_mean',0) or 0 for c in sem_campos]
        off = (vi-(n_ver-1)/2)*bw
        ax1.bar(x+off, f1s, bw*0.9, label=ver, color=vch(vi), alpha=0.85)
    ax1.axhline(0.70, color='#2E7D32', linestyle='--', linewidth=1.0, alpha=0.6, label='Highly Similar (0.70)')
    ax1.axhline(0.60, color='#8BC34A', linestyle='--', linewidth=1.0, alpha=0.6, label='Moderately Similar (0.60)')
    ax1.axhline(0.40, color='#F9A825', linestyle='--', linewidth=1.0, alpha=0.6, label='Slightly Similar (0.40)')
    ax1.set_xticks(x); ax1.set_xticklabels(sem_campos, rotation=30, ha='right', fontsize=11.0)
    ax1.set_ylim(0, 1.12); ax1.set_ylabel('Soft-F1 mean', fontsize=12.0)
    ax1.grid(axis='y', alpha=0.4); ax1.spines[['top','right']].set_visible(False)
    vd   = versions_data[lv]
    prec = [vd.get(c,{}).get('bertscore_proxy',{}).get('precision',0) or 0 for c in sem_campos]
    rec  = [vd.get(c,{}).get('bertscore_proxy',{}).get('recall',0)    or 0 for c in sem_campos]
    f1s  = [vd.get(c,{}).get('bertscore_proxy',{}).get('f1_mean',0)   or 0 for c in sem_campos]
    bw2  = 0.26; ax2 = axes[1]; ax2.set_facecolor('#F5F7FA')
    ax2.bar(x-bw2, prec, bw2, label='Precision', color=METRIC_COLORS['precision'], alpha=0.85)
    ax2.bar(x,     f1s,  bw2, label='F1',        color=METRIC_COLORS['bertscore_f1'], alpha=0.85)
    ax2.bar(x+bw2, rec,  bw2, label='Recall',    color=METRIC_COLORS['recall'], alpha=0.85)
    ax2.set_xticks(x); ax2.set_xticklabels(sem_campos, rotation=30, ha='right', fontsize=11.0)
    ax2.set_ylim(0, 1.12); ax2.set_ylabel('Score', fontsize=12.0)
    ax2.legend(fontsize=11.0, loc='upper left', bbox_to_anchor=(1.02, 1.0),
               borderaxespad=0, framealpha=0.9)
    ax2.grid(axis='y', alpha=0.4); ax2.spines[['top','right']].set_visible(False)
    buf = fig_to_buf(fig); plt.close('all')
    return buf, 800, 311


def chart_semantic_comparison(versions_data, versions, sem_campos, version_llms=None):
    version_llms = version_llms or {}
    llms_ordered, seen_llms = [], set()
    for v in versions:
        llm = version_llms.get(v, v.split('|')[0] if '|' in v else v)
        if llm not in seen_llms: llms_ordered.append(llm); seen_llms.add(llm)
    LLM_PAL = ['#1565C0','#2E7D32','#6A1B9A','#D84315','#00695C',
               '#4E342E','#37474F','#1A237E','#880E4F','#E65100']
    llm_color = {l: LLM_PAL[i % len(LLM_PAL)] for i, l in enumerate(llms_ordered)}
    ver_ls    = {1: '-', 2: '--', 3: ':'}

    n_ver = len(versions); x = np.arange(len(sem_campos))
    fig_h = max(n_ver * 0.28 + 5.5, 7.0)
    fig, axes = plt.subplots(1, 2, figsize=(18, fig_h))
    fig.subplots_adjust(wspace=0.46, left=0.05, right=0.70, top=0.90, bottom=0.22)

    ax1 = axes[0]
    for ver in versions:
        vd  = versions_data.get(ver, {})
        llm = version_llms.get(ver, ver.split('|')[0] if '|' in ver else ver)
        lbl = ver.split('|')[-1] if '|' in ver else ver
        vnum = int(lbl[1:]) if lbl.startswith('V') and lbl[1:].isdigit() else 1
        ax1.plot(x, [vd.get(c,{}).get('rouge_l',{}).get('mean',0) or 0 for c in sem_campos],
                 linestyle=ver_ls.get(vnum, '-'), color=llm_color.get(llm, '#333333'),
                 linewidth=1.4, markersize=3.5, alpha=0.85, label=f'{llm} {lbl}')
    ax1.axhline(0.70, color='#1A9641', linestyle='--', linewidth=0.9, alpha=0.65)
    ax1.axhline(0.60, color='#A6D96A', linestyle='--', linewidth=0.9, alpha=0.65)
    ax1.axhline(0.40, color='#FDAE61', linestyle='--', linewidth=0.9, alpha=0.65)
    ax1.set_xticks(x); ax1.set_xticklabels(sem_campos, rotation=32, ha='right')
    ax1.set_ylim(0, 1.05); ax1.set_ylabel('ROUGE-L (mean)')
    ax1.grid(axis='y', alpha=0.4, linewidth=0.5)
    ax1.spines[['top','right']].set_visible(False)

    ax2 = axes[1]
    plotted = set()
    for ver in versions:
        vd  = versions_data.get(ver, {})
        llm = version_llms.get(ver, ver.split('|')[0] if '|' in ver else ver)
        color = llm_color.get(llm, '#333333')
        rl  = [vd.get(c,{}).get('rouge_l',{}).get('mean',0) or 0 for c in sem_campos]
        sf1 = [vd.get(c,{}).get('bertscore_proxy',{}).get('f1_mean',0) or 0 for c in sem_campos]
        tf1 = [vd.get(c,{}).get('token_f1',{}).get('mean',0) or 0 for c in sem_campos]
        ax2.scatter(rl, sf1, s=[v*220+20 for v in tf1], color=color, alpha=0.60,
                    edgecolors='white', linewidths=0.6, label=llm if llm not in plotted else None)
        plotted.add(llm)
    ax2.plot([0,1],[0,1],'k--', alpha=0.18, linewidth=0.8)
    ax2.set_xlim(-0.02, 1.05); ax2.set_ylim(-0.02, 1.05)
    ax2.set_xlabel('ROUGE-L (mean)'); ax2.set_ylabel('Soft-F1 (mean)')
    ax2.legend(fontsize=11, loc='upper left', bbox_to_anchor=(1.02, 1.0), borderaxespad=0)
    ax2.grid(alpha=0.3, linewidth=0.5); ax2.spines[['top','right']].set_visible(False)

    buf = fig_to_buf(fig, dpi=180); plt.close('all')
    return buf, 800, 300


def chart_llm(llm_name, llm_versions, versions_data, sem_campos):
    det_campos = ['cvss', 'port', 'protocol', 'severity', 'references']
    det_labels = ['CVSS', 'Port', 'Protocol', 'Severity\n(F1)', 'References\n(Set-F1)']
    n_sem = len(sem_campos); n_det = len(det_campos); n_ver = len(llm_versions)
    bw    = min(0.70 / max(n_ver, 1), 0.26)
    offsets = np.linspace(-(n_ver-1)/2, (n_ver-1)/2, n_ver) * bw
    VER_PAL = ['#1565C0', '#2E7D32', '#6A1B9A']

    fig, axes = plt.subplots(1, 2, figsize=(16, 6))
    fig.subplots_adjust(wspace=0.42, left=0.06, right=0.72, top=0.88, bottom=0.22)

    for ax_idx, (ax, campo_list, ylbl, ttl) in enumerate(zip(
            axes,
            [sem_campos, det_campos],
            ['ROUGE-L (mean)', 'Score'],
            ['Semantic Fields', 'Deterministic Fields'])):
        ax.set_facecolor('white')
        x = np.arange(len(campo_list))
        for vi, (ver, off) in enumerate(zip(llm_versions, offsets)):
            vd    = versions_data.get(ver, {})
            lbl   = ver.split('|')[-1] if '|' in ver else ver
            color = VER_PAL[vi % len(VER_PAL)]
            if ax_idx == 0:
                vals = [vd.get(c, {}).get('rouge_l', {}).get('mean', 0) or 0 for c in campo_list]
            else:
                vals = []
                for c in campo_list:
                    r = vd.get(c, {})
                    if c == 'severity':     vals.append(r.get('f1_macro', 0) or 0)
                    elif c == 'references': vals.append(r.get('set_f1_cve', {}).get('mean', 0) or 0)
                    else:                   vals.append(r.get('exact_match_rate', 0) or 0)
            ax.bar(x + off, vals, bw * 0.90, label=lbl, color=color, alpha=0.85,
                   edgecolor='white', linewidth=0.4)
        import matplotlib.lines as mlines, matplotlib.patches as mpatches
        ax.axhline(0.70, color='#1A9641', linestyle='--', linewidth=0.9, alpha=0.75)
        ax.axhline(0.60, color='#A6D96A', linestyle='--', linewidth=0.9, alpha=0.75)
        ax.axhline(0.40, color='#FDAE61', linestyle='--', linewidth=0.9, alpha=0.75)
        xlabels = sem_campos if ax_idx == 0 else det_labels
        ax.set_xticks(x); ax.set_xticklabels(xlabels, rotation=32, ha='right')
        ax.set_ylim(0, 1.05); ax.set_ylabel(ylbl)
        ax.grid(axis='y', alpha=0.4, linewidth=0.5)
        ax.spines[['top', 'right']].set_visible(False)
        if ax_idx == 1:
            cat_handles = [
                mlines.Line2D([], [], color='#1A9641', linestyle='--', linewidth=1.2, label='Highly Similar (≥0.70)'),
                mlines.Line2D([], [], color='#A6D96A', linestyle='--', linewidth=1.2, label='Moderately Similar (0.60–0.70)'),
                mlines.Line2D([], [], color='#FDAE61', linestyle='--', linewidth=1.2, label='Slightly Similar (0.40–0.60)'),
                mpatches.Patch(color='#D7191C', alpha=0.55, label='Divergent (<0.40)'),
            ]
            ax.legend(handles=ax.get_legend_handles_labels()[0] + cat_handles,
                      fontsize=11, loc='upper left', bbox_to_anchor=(1.02, 1.0),
                      borderaxespad=0, framealpha=0.9)

    buf = fig_to_buf(fig, dpi=180); plt.close('all')
    return buf, 800, 300


def chart_set_f1(versions_data, versions):
    n_ver = len(versions)
    means = [versions_data[ver].get('references',{}).get('set_f1_cve',{}).get('mean',0) or 0 for ver in versions]
    mins  = [versions_data[ver].get('references',{}).get('set_f1_cve',{}).get('min',0)  or 0 for ver in versions]
    maxs  = [versions_data[ver].get('references',{}).get('set_f1_cve',{}).get('max',0)  or 0 for ver in versions]
    fig, axes = plt.subplots(1, 2, figsize=(16, 6), facecolor='#FAFAFA')
    fig.subplots_adjust(wspace=0.4, left=0.07, right=0.74, top=0.88, bottom=0.16)
    xv = np.arange(n_ver); ax1 = axes[0]; ax1.set_facecolor('#F5F7FA')
    ax1.bar(xv, means, 0.5, color=[vch(i) for i in range(n_ver)], alpha=0.85,
            yerr=[[m-mn for m,mn in zip(means,mins)],[mx-m for m,mx in zip(means,maxs)]],
            capsize=6, error_kw={'elinewidth':1.5,'ecolor':'#555555'})
    ax1.axhline(0.70, color='#2E7D32', linestyle='--', linewidth=1.0, alpha=0.6, label='Highly Similar (0.70)')
    ax1.axhline(0.60, color='#8BC34A', linestyle='--', linewidth=1.0, alpha=0.6, label='Moderately Similar (0.60)')
    ax1.axhline(0.40, color='#F9A825', linestyle='--', linewidth=1.0, alpha=0.6, label='Slightly Similar (0.40)')
    ax1.set_xticks(xv); ax1.set_xticklabels(versions, fontsize=11.0, rotation=30, ha='right')
    ax1.set_ylim(0, 1.12)
    ax1.set_ylabel('Set-F1 CVE (mean ± min/max)', fontsize=12.0)
    ax1.legend(fontsize=11.0, loc='lower right', framealpha=0.9)
    ax1.grid(axis='y', alpha=0.4); ax1.spines[['top','right']].set_visible(False)
    ax2 = axes[1]; ax2.remove(); ax2 = fig.add_subplot(1, 2, 2)
    ax2.set_facecolor('#F5F7FA'); ax2.set_xlim(0, 1); ax2.set_ylim(0, 1); ax2.axis('off')
    lm = means[-1]; lv = versions[-1]
    color = '#2E7D32' if lm>=0.70 else ('#8BC34A' if lm>=0.60 else ('#F9A825' if lm>=0.40 else '#D32F2F'))
    ax2.add_patch(plt.Circle((0.5,0.5),0.38,color='#EEEEEE',zorder=1))
    ax2.add_patch(plt.Circle((0.5,0.5),0.38,color=color,alpha=0.15,zorder=2))
    ax2.text(0.5,0.54,f'{lm:.4f}',ha='center',va='center',fontsize=30.0,fontweight='bold',color=color,zorder=3)
    ax2.text(0.5,0.38,'Set-F1 CVE',ha='center',va='center',fontsize=13.0,color='#555555',zorder=3)
    ax2.text(0.5,0.30,lv,ha='center',va='center',fontsize=14.0,fontweight='bold',color='#1F3864',zorder=3)
    buf = fig_to_buf(fig); plt.close('all')
    return buf, 800, 300


def chart_exact_match(versions_data, versions):
    det_c = ['cvss','port','protocol']; n_ver = len(versions); bw = 0.8/max(n_ver,1)
    xd = np.arange(len(det_c))
    fig, axes = plt.subplots(1, 2, figsize=(16, 6), facecolor='#FAFAFA')
    fig.subplots_adjust(wspace=0.4, left=0.07, right=0.74, top=0.88, bottom=0.14)
    ax1 = axes[0]; ax1.set_facecolor('#F5F7FA')
    for vi, ver in enumerate(versions):
        vd = versions_data[ver]
        vals = [vd.get(c,{}).get('exact_match_rate',0) or 0 for c in det_c]
        off  = (vi-(n_ver-1)/2)*bw
        ax1.bar(xd+off, vals, bw*0.9, label=ver, color=vch(vi), alpha=0.85)
    all_vals = [versions_data[ver].get(c,{}).get('exact_match_rate',0) or 0
                for ver in versions for c in det_c]
    y_min = max(0, min(all_vals) - 0.05)
    ax1.axhline(0.99, color='#2E7D32', linestyle='--', linewidth=1.2, alpha=0.7, label='≥ 0.99')
    ax1.set_xticks(xd); ax1.set_xticklabels(det_c, fontsize=13.0); ax1.set_ylim(y_min, 1.015)
    ax1.set_ylabel('Exact Match Rate', fontsize=12.0)
    ax1.legend(fontsize=11.0, loc='lower right', framealpha=0.9)
    ax1.grid(axis='y', alpha=0.4); ax1.spines[['top','right']].set_visible(False)
    ax2 = axes[1]; ax2.set_facecolor('#F5F7FA'); ax2.axis('off')
    lv = versions[-1]; vd = versions_data[lv]
    for i, campo in enumerate(det_c):
        val = vd.get(campo,{}).get('exact_match_rate',0) or 0; cx = 0.18+i*0.32
        color = '#2E7D32' if val>=0.99 else ('#F9A825' if val>=0.95 else '#D32F2F')
        ax2.add_patch(plt.Circle((cx,0.55),0.14,color='#EEEEEE',transform=ax2.transAxes,zorder=1))
        ax2.add_patch(plt.Circle((cx,0.55),0.14,color=color,alpha=0.18,transform=ax2.transAxes,zorder=2))
        ax2.text(cx,0.57,f'{val:.4f}',ha='center',va='center',fontsize=16.0,fontweight='bold',color=color,transform=ax2.transAxes)
        ax2.text(cx,0.43,campo,ha='center',va='center',fontsize=12.0,color='#444444',transform=ax2.transAxes)
    ax2.text(0.5,0.88,f'Exact Match — {lv}',ha='center',va='center',fontsize=14.0,fontweight='bold',color='#1F3864',transform=ax2.transAxes)
    buf = fig_to_buf(fig); plt.close('all')
    return buf, 800, 300


def chart_severity(versions_data, versions):
    lv = versions[-1]; sev_r = versions_data[lv].get('severity',{})
    cm_data = sev_r.get('confusion_matrix',{}); labels = cm_data.get('labels',[]); matrix = cm_data.get('matrix',[])
    pc = cm_data.get('per_class',{})
    fig = plt.figure(figsize=(18, 7), facecolor='#FAFAFA')
    gs = GridSpec(1, 3, figure=fig, wspace=0.50, left=0.06, right=0.84, top=0.88, bottom=0.14)
    ax1 = fig.add_subplot(gs[0,0]); ax1.set_facecolor('#F5F7FA')
    ver_f1 = [(ver,
               versions_data[ver].get('severity',{}).get('f1_macro',0) or 0,
               versions_data[ver].get('severity',{}).get('precision_macro',0) or 0,
               versions_data[ver].get('severity',{}).get('recall_macro',0) or 0)
              for ver in versions]
    xvf = np.arange(len(ver_f1))
    ax1.bar(xvf-0.25,[v[2] for v in ver_f1],0.25,label='Precision',color=METRIC_COLORS['precision'],alpha=0.85)
    ax1.bar(xvf,     [v[1] for v in ver_f1],0.25,label='F1-macro', color=METRIC_COLORS['f1_macro'],alpha=0.85)
    ax1.bar(xvf+0.25,[v[3] for v in ver_f1],0.25,label='Recall',   color=METRIC_COLORS['recall'],alpha=0.85)
    ax1.set_xticks(xvf); ax1.set_xticklabels([v[0] for v in ver_f1], fontsize=9.0, rotation=30, ha='right')
    ax1.set_ylim(0, 1.12); ax1.set_ylabel('Macro score', fontsize=12.0)
    ax1.legend(fontsize=11.0, loc='lower right', framealpha=0.9)
    ax1.grid(axis='y', alpha=0.4); ax1.spines[['top','right']].set_visible(False)
    ax2 = fig.add_subplot(gs[0,1]); ax2.set_facecolor('#F5F7FA')
    if pc:
        classes = list(pc.keys()); f1_vals = [[pc[c].get('f1-score',0)] for c in classes]
        discrete_heatmap(ax2, f1_vals, ['F1-score'], classes,
                         color_fn=lambda i,j,v: score_bg(v),
                         val_fmt=lambda v: f'{v:.3f}' if v is not None else 'N/A', fontsize=11.0)
        ax2.set_xticks([0.5]); ax2.set_xticklabels(['F1-score'], fontsize=11.0)
    ax3 = fig.add_subplot(gs[0,2]); ax3.set_facecolor('#F5F7FA')
    if matrix:
        def cm_color(i,j,v): return '#C8E6C9' if i==j else ('#FFCCBC' if v>0 else '#F5F5F5')
        discrete_heatmap(ax3, matrix, labels, labels, color_fn=cm_color,
                         val_fmt=lambda v: str(int(v)) if v is not None else 'N/A', fontsize=11.0)
        ax3.set_xlabel('Predicted (merged)', fontsize=11.0, labelpad=6)
        ax3.set_ylabel('Gold (baseline)', fontsize=11.0, labelpad=6)
    buf = fig_to_buf(fig); plt.close('all')
    return buf, 800, 311


# ═══════════════════════════════════════════════════════════════════════════
# XLSX BUILDER
# ═══════════════════════════════════════════════════════════════════════════

def chart_omission_hallucination(versions, version_dfs, baseline_df, version_llms=None):
    """
    Two subplots: Omission (top) and Hallucination (bottom).
    X = fields.  Color = LLM.  Line style = version (V1 solid, V2 dashed, V3 dotted).
    """
    version_llms  = version_llms or {}
    mapped_fields = {c: FIELD_MAP[c] for c in CAMPO_ORDER
                     if FIELD_MAP[c][0] is not None and FIELD_MAP[c][1] is not None}
    fields   = list(mapped_fields.keys())
    n_fields = len(fields)

    def _ie(val):
        if val is None: return True
        return str(val).strip() in ('', 'nan', 'None', '[]', "['']", 'NaN')

    bl_clean = baseline_df.copy()
    bl_clean.columns = [c.strip().lstrip('﻿') for c in bl_clean.columns]
    bl_clean['_ip_key']   = bl_clean['IP'].astype(str).str.strip()
    bl_clean['_name_key'] = bl_clean['NVT Name'].astype(str).str.strip()
    bl_names = bl_clean['_name_key'].unique()

    omit_rates, halluc_rates = {}, {}
    for ver in versions:
        if ver not in version_dfs:
            continue
        merged = version_dfs[ver].copy()
        merged.columns = [c.strip().lstrip('﻿') for c in merged.columns]
        # Name mapping ≥85% + exact host
        mg_names = merged['Name'].dropna().astype(str).str.strip().unique()
        nm = _build_name_map(mg_names, bl_names, threshold=0.85)
        merged['_mapped_name'] = merged['Name'].astype(str).str.strip().map(nm)
        merged['_host_key']    = merged['host'].astype(str).str.strip()
        mg_dedup = (merged.dropna(subset=['_mapped_name'])
                          .sort_values(['_host_key', '_mapped_name'])
                          .groupby(['_host_key', '_mapped_name'], sort=False)
                          .first().reset_index())
        bl_dedup = (bl_clean.sort_values(['_ip_key', '_name_key'])
                            .groupby(['_ip_key', '_name_key'], sort=False)
                            .first().reset_index())
        df = mg_dedup.merge(bl_dedup,
                            left_on=['_host_key', '_mapped_name'],
                            right_on=['_ip_key',  '_name_key'],
                            how='inner', suffixes=('_mg', '_bl'))
        n_total = len(df)
        o_r, h_r = [], []
        for campo in fields:
            col_mg, col_bl = mapped_fields[campo]
            mg_col = col_mg + '_mg' if col_mg + '_mg' in df.columns else col_mg
            bl_col = col_bl + '_bl' if col_bl + '_bl' in df.columns else col_bl
            if mg_col in df.columns:
                mg_e = df[mg_col].apply(_is_empty)
            else:
                mg_e = pd.Series([True] * n_total, index=df.index)
            if bl_col in df.columns:
                bl_e = df[bl_col].apply(_is_empty)
            else:
                bl_e = pd.Series([True] * n_total, index=df.index)
            o_r.append((mg_e & ~bl_e).sum() / n_total * 100 if n_total else 0)
            h_r.append((~mg_e & bl_e).sum()  / n_total * 100 if n_total else 0)
        omit_rates[ver], halluc_rates[ver] = o_r, h_r

    LLM_PAL = {'deepseek': '#1565C0', 'gpt4': '#6A1B9A', 'gpt5': '#2E7D32',
               'llama3': '#BF360C',  'llama4': '#006064'}
    FALLBACK = ['#1565C0','#6A1B9A','#2E7D32','#BF360C','#006064','#D84315','#0277BD']
    VER_MK   = {'TMMv1': ('o', '-'), 'TMMv2': ('s', '--'), 'TMMv3': ('D', ':')}

    llm_ordered, seen_llm = [], set()
    for ver in versions:
        llm = version_llms.get(ver, ver.split('|')[0] if '|' in ver else ver)
        if llm not in seen_llm:
            llm_ordered.append(llm); seen_llm.add(llm)
    for i, llm in enumerate(llm_ordered):
        if llm not in LLM_PAL:
            LLM_PAL[llm] = FALLBACK[i % len(FALLBACK)]

    # Threshold unificado: uma única linha/rótulo "20% threshold" vale para os
    # dois painéis (omissão e alucinação), evitando legenda repetida.
    THRESH_COLOR = '#E65100'

    x = np.arange(n_fields)
    if PAPER_LAYOUT:
        # Figura-fonte mais estreita => escala menos ao encaixar no paper =>
        # fontes maiores no PDF final. Legenda única à direita.
        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(11.5, 9.8),
                                       gridspec_kw={'hspace': 0.95})
        fig.subplots_adjust(left=0.09, right=0.70, top=0.94, bottom=0.14)
    else:
        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(18, 12),
                                       gridspec_kw={'hspace': 0.50})
        fig.subplots_adjust(left=0.05, right=0.72, top=0.93, bottom=0.10)

    _lab_fs = 16 if PAPER_LAYOUT else 11
    _ttl_fs = 17 if PAPER_LAYOUT else 12

    def _draw(ax, rates_dict, title):
        ax.set_facecolor('white')
        ax.axhline(0, color='#CCCCCC', linewidth=0.7)
        for ver in versions:
            if ver not in rates_dict: continue
            llm    = version_llms.get(ver, ver.split('|')[0] if '|' in ver else ver)
            vlabel = ver.split('|')[-1] if '|' in ver else ver
            mk, ls = VER_MK.get(vlabel, ('o', '-'))
            color  = LLM_PAL.get(llm, '#444444')
            ax.plot(x, rates_dict[ver], color=color, linewidth=1.8,
                    linestyle=ls, alpha=0.80, zorder=3)
            ax.scatter(x, rates_dict[ver], s=85, marker=mk, color=color,
                       zorder=5, edgecolors='white', linewidths=0.8, alpha=0.95)
        ax.axhline(20, color=THRESH_COLOR, linestyle='--', linewidth=1.4, alpha=0.80, zorder=6)
        ax.set_xticks(x)
        ax.set_xticklabels(fields, rotation=35, ha='right', fontsize=_lab_fs)
        ax.tick_params(axis='y', labelsize=_lab_fs - 1)
        ax.margins(x=0.03)
        ax.set_ylabel('Rate (%)', fontsize=_lab_fs)
        ax.set_title(title, fontsize=_ttl_fs, fontweight='bold', pad=10)
        y_max = max((max(v) for v in rates_dict.values() if v), default=0)
        ax.set_ylim(-2, max(y_max + 10, 30))
        ax.grid(axis='y', alpha=0.30, linewidth=0.5)
        ax.spines[['top', 'right']].set_visible(False)
        for xi in range(n_fields):
            if any(rates_dict.get(v, [0]*n_fields)[xi] > 20 for v in versions):
                ax.axvspan(xi - 0.45, xi + 0.45, alpha=0.06, color=THRESH_COLOR, zorder=0)

    _draw(ax1, omit_rates,   'Omission Rate  (merged empty, baseline filled)')
    _draw(ax2, halluc_rates, 'Hallucination Rate  (merged filled, baseline empty)')

    # Uma única legenda compartilhada pelos dois painéis, centralizada à direita.
    import matplotlib.lines as mlines, matplotlib.patches as mpatches
    llm_h = [mpatches.Patch(color=LLM_PAL[llm], label=llm) for llm in llm_ordered]
    ver_h = [
        mlines.Line2D([], [], color='#555', linestyle='-',  marker='o', ms=8, label='TMMv1'),
        mlines.Line2D([], [], color='#555', linestyle='--', marker='s', ms=8, label='TMMv2'),
        mlines.Line2D([], [], color='#555', linestyle=':',  marker='D', ms=8, label='TMMv3'),
        mlines.Line2D([], [], color=THRESH_COLOR, linestyle='--', linewidth=1.4, label='20% threshold'),
    ]
    fig.legend(handles=llm_h + ver_h, fontsize=(16 if PAPER_LAYOUT else 11),
               loc='center left', bbox_to_anchor=(0.715, 0.5),
               borderaxespad=0, framealpha=0.92, edgecolor='#CCCCCC')

    buf = fig_to_buf(fig); plt.close('all')
    return buf, 920, 580
_EMPTY_TOKENS = frozenset({
    '', 'nan', 'none', 'null', 'n/a', 'na', '-', '[]', "['']",
    '[""]', '[none]', '[null]', 'NaN', 'None', 'NULL', 'N/A',
})

def _is_empty(val):
    """Return True if val is considered empty/null (case-insensitive, whitespace-stripped)."""
    if val is None:
        return True
    s = str(val).strip()
    return s.lower() in _EMPTY_TOKENS or s in _EMPTY_TOKENS


def _build_omission_matrix(wb, versions, version_dfs, baseline_df, version_llms=None):
    """
    Omission Summary sheet — rows = (LLM, Version), columns = aggregate OK/Omission/Halluc/NA rates.
    The scatter chart shows X=omission%, Y=hallucination%, one point per (LLM, Version),
    aggregated across all mapped fields.
    """
    version_llms  = version_llms or {}
    mapped_fields = {c: FIELD_MAP[c] for c in CAMPO_ORDER
                     if FIELD_MAP[c][0] is not None and FIELD_MAP[c][1] is not None}
    fields = list(mapped_fields.keys())

    STATUS_OK, STATUS_O, STATUS_A, STATUS_NA = 'OK', 'O', 'A', 'N/A'
    COLOR_OK = 'C8E6C9'; COLOR_O = 'FFCCBC'; COLOR_A = 'FFF9C4'; COLOR_NA = 'EEEEEE'

    LLM_PAL = {'deepseek': '#1565C0', 'gpt4': '#6A1B9A', 'gpt5': '#2E7D32',
               'llama3': '#BF360C',  'llama4': '#006064'}
    FALLBACK = ['#1565C0','#6A1B9A','#2E7D32','#BF360C','#006064','#D84315']
    llm_ordered, seen_llm = [], set()
    for ver in versions:
        llm = version_llms.get(ver, ver.split('|')[0] if '|' in ver else ver)
        if llm not in seen_llm:
            llm_ordered.append(llm); seen_llm.add(llm)
    for i, llm in enumerate(llm_ordered):
        if llm not in LLM_PAL:
            LLM_PAL[llm] = FALLBACK[i % len(FALLBACK)]

    # Compute per-version aggregate counts — host+name merge, vectorized
    bl_clean = baseline_df.copy()
    bl_clean.columns = [c.strip().lstrip('﻿') for c in bl_clean.columns]
    bl_clean['_ip_key']   = bl_clean['IP'].astype(str).str.strip()
    bl_clean['_name_key'] = bl_clean['NVT Name'].astype(str).str.strip()
    bl_names_om = bl_clean['_name_key'].unique()
    bl_dedup_om = (bl_clean.sort_values(['_ip_key', '_name_key'])
                           .groupby(['_ip_key', '_name_key'], sort=False)
                           .first().reset_index())
    agg = {}   # ver -> {'ok':n, 'o':n, 'a':n, 'na':n, 'total_obs':n, 'n_pairs':n}
    for ver in versions:
        if ver not in version_dfs:
            continue
        merged = version_dfs[ver].copy()
        merged.columns = [c.strip().lstrip('﻿') for c in merged.columns]
        mg_names_om = merged['Name'].dropna().astype(str).str.strip().unique()
        nm_om = _build_name_map(mg_names_om, bl_names_om, threshold=0.85)
        merged['_mapped_name'] = merged['Name'].astype(str).str.strip().map(nm_om)
        merged['_host_key']    = merged['host'].astype(str).str.strip()
        mg_dedup_om = (merged.dropna(subset=['_mapped_name'])
                              .sort_values(['_host_key', '_mapped_name'])
                              .groupby(['_host_key', '_mapped_name'], sort=False)
                              .first().reset_index())
        df = mg_dedup_om.merge(bl_dedup_om,
                               left_on=['_host_key', '_mapped_name'],
                               right_on=['_ip_key',  '_name_key'],
                               how='inner', suffixes=('_mg', '_bl'))
        n_pairs = len(df)
        ok = o = a = na = 0
        for campo in fields:
            col_mg, col_bl = mapped_fields[campo]
            mg_col = col_mg + '_mg' if col_mg + '_mg' in df.columns else col_mg
            bl_col = col_bl + '_bl' if col_bl + '_bl' in df.columns else col_bl
            # Vectorized empty detection
            if mg_col in df.columns:
                mg_e = df[mg_col].apply(_is_empty)
            else:
                mg_e = pd.Series([True] * n_pairs, index=df.index)
            if bl_col in df.columns:
                bl_e = df[bl_col].apply(_is_empty)
            else:
                bl_e = pd.Series([True] * n_pairs, index=df.index)
            ok += int((~mg_e & ~bl_e).sum())
            o  += int(( mg_e & ~bl_e).sum())
            a  += int((~mg_e &  bl_e).sum())
            na += int(( mg_e &  bl_e).sum())
        total_obs = n_pairs * len(fields)
        agg[ver] = {'ok': ok, 'o': o, 'a': a, 'na': na,
                    'total_obs': total_obs, 'n_pairs': n_pairs}

    # ── Excel sheet ──────────────────────────────────────────────────────────
    ws = wb.create_sheet('Omission Summary')
    ws.merge_cells('A1:J1')
    c = ws['A1']
    c.value = 'Omission & Hallucination — Cross-Version Summary'
    c.font  = Font(bold=True, color='FFFFFF', name='Arial', size=16)
    c.fill  = PatternFill('solid', start_color=PALETTE['dark'])
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 34

    hdr = ['LLM', 'Version', 'N Pairs', 'N Field Obs.',
           'OK', 'OK %', 'Omission', 'Omission %', 'Hallucination', 'Halluc. %']
    for ci, h in enumerate(hdr, 1):
        bc(ws, 2, ci, h, bold=True, fg='FFFFFF', bg=PALETTE['dark'], size=11)
    ws.row_dimensions[2].height = 24

    row_num = 3
    for ver in versions:
        if ver not in agg:
            continue
        llm    = version_llms.get(ver, ver.split('|')[0] if '|' in ver else ver)
        vlabel = ver.split('|')[-1] if '|' in ver else ver
        d      = agg[ver]
        tot    = d['total_obs'] or 1
        pct_ok = round(d['ok'] / tot * 100, 1)
        pct_o  = round(d['o']  / tot * 100, 1)
        pct_a  = round(d['a']  / tot * 100, 1)
        bg     = PALETTE['white'] if row_num % 2 == 0 else PALETTE['alt']

        vals = [llm, vlabel, d['n_pairs'], d['total_obs'],
                d['ok'], f'{pct_ok}%', d['o'], f'{pct_o}%', d['a'], f'{pct_a}%']
        bgs  = [bg, bg, bg, bg,
                COLOR_OK, COLOR_OK, COLOR_O, COLOR_O, COLOR_A, COLOR_A]
        for ci, (val, cbg) in enumerate(zip(vals, bgs), 1):
            cell = ws.cell(row=row_num, column=ci, value=val)
            cell.font      = Font(bold=(ci in (1, 2)), name='Arial', size=10)
            cell.fill      = PatternFill('solid', start_color=cbg)
            cell.alignment = Alignment(horizontal='center' if ci > 2 else 'left',
                                       vertical='center')
            cell.border    = BORDER_THIN
        ws.row_dimensions[row_num].height = 20
        row_num += 1

    col_widths = [14, 10, 10, 14, 10, 10, 12, 12, 16, 12]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A3'

    # ── Scatter: Omission% × Hallucination% per LLM × Version ────────────
    import matplotlib.pyplot as plt, matplotlib.patches as mpatches
    import matplotlib.lines as mlines
    from openpyxl.drawing.image import Image as XLImage

    VER_MK = {'TMMv1': ('o', 160), 'TMMv2': ('s', 160), 'TMMv3': ('D', 160)}

    X_MAX = 21; Y_MAX = 8
    if PAPER_LAYOUT:
        # Mais alto e estreito, legenda embaixo: casa com o painel de linhas
        # (omission_halluc) quando usados lado a lado no paper.
        fig, ax = plt.subplots(figsize=(6.8, 7.2))
        fig.subplots_adjust(left=0.12, right=0.97, top=0.97, bottom=0.24)
    else:
        fig, ax = plt.subplots(figsize=(10, 6))
        fig.subplots_adjust(left=0.10, right=0.72, top=0.88, bottom=0.13)
    ax.set_facecolor('white')
    ax.set_xlabel('Omission Rate (%), avg. across fields',  fontsize=14, labelpad=6)
    ax.set_ylabel('Hallucination Rate (%), avg. across fields', fontsize=14, labelpad=6)
    ax.tick_params(axis='both', labelsize=13)

    for ver in versions:
        if ver not in agg: continue
        llm    = version_llms.get(ver, ver.split('|')[0] if '|' in ver else ver)
        vlabel = ver.split('|')[-1] if '|' in ver else ver
        d      = agg[ver]
        tot    = d['total_obs'] or 1
        px     = d['o'] / tot * 100
        py     = d['a'] / tot * 100
        color  = LLM_PAL.get(llm, '#444444')
        mk, ms = VER_MK.get(vlabel, ('o', 140))
        ax.scatter(px, py, s=ms, marker=mk, color=color,
                   edgecolors='white', linewidths=1.2, alpha=0.92, zorder=5)

    ax.axhline(0, color='#BBBBBB', linewidth=0.8)
    ax.axvline(0, color='#BBBBBB', linewidth=0.8)
    ax.fill_between([20, X_MAX], 0, Y_MAX, color='#D32F2F', alpha=0.06, zorder=0)
    ax.axvline(20, color='#E65100', linewidth=1.2, linestyle='--', alpha=0.80)
    ax.set_xlim(0, X_MAX); ax.set_ylim(0, Y_MAX)
    ax.set_xticks(range(0, X_MAX + 1, 2))
    ax.set_yticks(range(0, Y_MAX + 1, 1))
    ax.grid(alpha=0.25, linewidth=0.6); ax.spines[['top', 'right']].set_visible(False)

    llm_h = [mpatches.Patch(color=LLM_PAL.get(l, '#444'), label=l) for l in llm_ordered]
    ver_h = [
        mlines.Line2D([], [], color='#555', marker='o', ls='none', ms=8, label='TMMv1'),
        mlines.Line2D([], [], color='#555', marker='s', ls='none', ms=8, label='TMMv2'),
        mlines.Line2D([], [], color='#555', marker='D', ls='none', ms=8, label='TMMv3'),
        mlines.Line2D([], [], color='#E65100', ls='--', lw=1.2, label='Omission > 20%'),
    ]
    if PAPER_LAYOUT:
        ax.legend(handles=llm_h + ver_h, fontsize=12, loc='upper center',
                  bbox_to_anchor=(0.5, -0.11), ncol=3, borderaxespad=0,
                  framealpha=0.92, columnspacing=1.2, handletextpad=0.5)
    else:
        ax.legend(handles=llm_h + ver_h, fontsize=10, loc='upper left',
                  bbox_to_anchor=(1.02, 1.0), borderaxespad=0, framealpha=0.92)

    export_vector(fig, 'omission_summary')
    import io
    buf_sc = io.BytesIO()
    fig.savefig(buf_sc, format='png', dpi=150, bbox_inches='tight')
    buf_sc.seek(0); plt.close('all')

    img_sc = XLImage(buf_sc)
    img_sc.width = 780; img_sc.height = 480
    chart_row = row_num + 2
    ws.add_image(img_sc, f'A{chart_row}')
    for r in range(chart_row, chart_row + 34):
        ws.row_dimensions[r].height = 15


def chart_all_llms(versions_data, versions, sem_campos, version_llms=None):
    """
    Grouped bar chart comparing all LLMs × versions side by side.
    Left: mean ROUGE-L across all semantic fields per LLM group (3 bars V1/V2/V3, error = std across fields).
    Right: mean score across all deterministic fields, same layout.
    """
    from collections import OrderedDict
    version_llms = version_llms or {}

    det_campos = ['cvss', 'port', 'protocol', 'severity', 'references']

    llm_groups: 'OrderedDict[str, list]' = OrderedDict()
    for ver in versions:
        llm = version_llms.get(ver, ver.split('|')[0] if '|' in ver else ver)
        llm_groups.setdefault(llm, []).append(ver)
    llm_names = list(llm_groups.keys())
    n_llm = len(llm_names)

    all_ver_labels = sorted({
        (ver.split('|')[-1] if '|' in ver else ver)
        for ver in versions
    })
    n_vl = len(all_ver_labels)

    VER_PAL = ['#1565C0', '#2E7D32', '#6A1B9A', '#D84315', '#006064']

    # ROUGE-L bands (Lin 2004 / NLP summarization literature)
    THRESH_SEM = [
        (0.70, '#1A9641', 'Highly Similar (>=0.70)'),
        (0.60, '#A6D96A', 'Mod. Similar (0.60-0.70)'),
        (0.40, '#FDAE61', 'Slightly Similar (0.40-0.60)'),
    ]
    # EM / F1-macro / Set-F1 bands (Rajpurkar 2016; Devlin 2018; general ML literature)
    THRESH_DET = [
        (0.90, '#1A9641', 'Excellent (>=0.90)'),
        (0.80, '#66BB6A', 'Good (0.80-0.90)'),
        (0.70, '#A6D96A', 'Decent (0.70-0.80)'),
        (0.50, '#FDAE61', 'Acceptable (0.50-0.70)'),
    ]

    bw  = min(0.70 / n_vl, 0.28)
    x   = np.arange(n_llm)
    if PAPER_LAYOUT:
        # Empilhado (2x1): proporção próxima do quadrado para uso lado a lado
        # com o heatmap Overview em meia coluna, sem esmagar os painéis.
        fig, axes = plt.subplots(2, 1, figsize=(9.2, 9.6))
        fig.subplots_adjust(hspace=0.42, left=0.10, right=0.77, top=0.95, bottom=0.07)
    else:
        fig, axes = plt.subplots(1, 2, figsize=(20, 6))
        fig.subplots_adjust(wspace=0.62, left=0.06, right=0.86, top=0.88, bottom=0.16)

    panels = [
        (axes[0], sem_campos, 'rouge_l',
         'Semantic Fields: ROUGE-L by LLM & Version',
         'ROUGE-L (mean +/- std across fields)', THRESH_SEM,
         'Divergent (<0.40)'),
        (axes[1], det_campos, 'det',
         'Deterministic Fields: EM / F1-macro / Set-F1 by LLM & Version',
         'Score (mean +/- std across fields)', THRESH_DET,
         'Poor (<0.50)'),
    ]

    for ax, field_list, metric_key, title, ylabel, thresholds, below_lbl in panels:
        import matplotlib.lines as mlines, matplotlib.patches as mpatches
        ax.set_facecolor('white')
        for vi, ver_label in enumerate(all_ver_labels):
            color = VER_PAL[vi % len(VER_PAL)]
            off   = (vi - n_vl / 2 + 0.5) * bw
            bar_vals, bar_errs = [], []
            for llm in llm_names:
                ver_key = next(
                    (v for v in llm_groups[llm]
                     if (v.split('|')[-1] if '|' in v else v) == ver_label), None)
                if ver_key is None:
                    bar_vals.append(0); bar_errs.append(0); continue
                vd = versions_data.get(ver_key, {})
                if metric_key == 'rouge_l':
                    vals = [vd.get(c, {}).get('rouge_l', {}).get('mean') or 0 for c in field_list]
                else:
                    vals = []
                    for c in field_list:
                        r = vd.get(c, {})
                        if   c == 'severity':   vals.append(r.get('f1_macro') or 0)
                        elif c == 'references': vals.append((r.get('set_f1_cve') or {}).get('mean') or 0)
                        else:                   vals.append(r.get('exact_match_rate') or 0)
                m = sum(vals) / len(vals) if vals else 0
                s = (sum((v - m) ** 2 for v in vals) / len(vals)) ** 0.5 if vals else 0
                bar_vals.append(m); bar_errs.append(s)
            ax.bar(x + off, bar_vals, bw * 0.90, label=ver_label, color=color, alpha=0.85,
                   edgecolor='white', linewidth=0.4, yerr=bar_errs, capsize=3,
                   error_kw={'elinewidth': 0.9, 'ecolor': '#333333', 'alpha': 0.7})
        for thr, col, lbl in thresholds:
            ax.axhline(thr, color=col, linestyle='--', linewidth=1.1, alpha=0.82, zorder=5)
        ax.set_xticks(x)
        ax.set_xticklabels(llm_names, fontsize=13, fontweight='bold')
        ax.tick_params(axis='y', labelsize=12)
        ax.set_ylim(0, 1.12); ax.set_ylabel(ylabel, fontsize=12)
        ax.set_title(title, fontsize=14, fontweight='bold', pad=10)
        ax.grid(axis='y', alpha=0.35, linewidth=0.5)
        ax.spines[['top', 'right']].set_visible(False)
        ver_handles = ax.get_legend_handles_labels()[0]
        sep = mpatches.Patch(visible=False, label='── Quality Bands ──')
        thr_handles = [
            mlines.Line2D([], [], color=col, linestyle='--', lw=1.4, label=lbl)
            for _, col, lbl in thresholds
        ]
        below_patch = mpatches.Patch(color='#D7191C', alpha=0.55, label=below_lbl)
        ax.legend(handles=ver_handles + [sep] + thr_handles + [below_patch],
                  fontsize=11, title='Version', title_fontsize=11,
                  loc='upper left', bbox_to_anchor=(1.02, 1.0),
                  borderaxespad=0, framealpha=0.92, edgecolor='#CCCCCC')
        for xi in np.arange(0.5, n_llm, 1):
            ax.axvline(xi, color='#CCCCCC', linewidth=0.7, zorder=0)

    buf = fig_to_buf(fig, dpi=180); plt.close('all')
    return buf, 900, 300


def build_xlsx(versions_data, versions, baseline_path, xlsx_path, version_pairs=None, version_llms=None):
    all_campos = [c for c in CAMPO_ORDER if CAMPO_META.get(c,{}).get('tipo') != 'N/A']
    sem_campos = [c for c in CAMPO_ORDER
                  if CAMPO_META.get(c,{}).get('tipo') == 'Semântico' and c != 'references']
    version_llms = version_llms or {}
    subtitle   = f'TMM — {", ".join(versions)} | {Path(baseline_path).name}'

    # Load raw DataFrames for omission/hallucination analysis
    def _strip_df_quotes(df):
        """Remove surrounding apostrophes from string values (CSV serialization artifact)."""
        def _sq(val):
            if isinstance(val, str):
                s = val.strip()
                if len(s) >= 2 and s[0] == "'" and s[-1] == "'":
                    return s[1:-1]
            return val
        for col in df.columns:
            if df[col].dtype == object:
                df[col] = df[col].apply(_sq)
        return df

    baseline_df  = pd.read_csv(baseline_path)
    version_dfs  = {}
    if version_pairs:
        for label, path, llm in version_pairs:
            key = f'{llm}|{label}' if llm else label
            version_dfs[key] = _strip_df_quotes(pd.read_csv(path))

    wb = Workbook()

    # ── Summary ──────────────────────────────────────────────────────────
    print('  Building Summary...')
    ws = wb.active; ws.title = 'Summary'
    n_ver = len(versions); total_cols = 3 + n_ver*3

    ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
    c = ws['A1']; c.value = 'TMM — Metrics Comparison by Version'
    c.font = Font(bold=True, color='FFFFFF', name='Arial', size=16)
    c.fill = PatternFill('solid', start_color=PALETTE['dark'])
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 36

    ws.merge_cells(f'A2:{get_column_letter(total_cols)}2')
    c2 = ws['A2']
    c2.value = (f'Baseline: {Path(baseline_path).name}  |  Versions: {", ".join(versions)}'
                f'  |  Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}')
    c2.font = Font(italic=True, color='444444', name='Arial', size=12)
    c2.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18

    for vi, ver in enumerate(versions):
        sc = 4+vi*3; ec = sc+2
        ws.merge_cells(f'{get_column_letter(sc)}3:{get_column_letter(ec)}3')
        c = ws.cell(row=3, column=sc,
                    value=f'{ver} — Version {ver[1:] if ver.startswith("V") else ver}')
        c.font = Font(bold=True, color='FFFFFF', name='Arial', size=13)
        c.fill = PatternFill('solid', start_color=vc(vi))
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER_THIN
    ws.row_dimensions[3].height = 26

    for ci, h in enumerate(['Field', 'Category', 'Main\nMetric'], 1):
        bc(ws, 4, ci, h, bold=True, fg='FFFFFF', bg=PALETTE['dark'], size=11)
    for vi, ver in enumerate(versions):
        sc = 4+vi*3; v = vc(vi)
        bc(ws, 4, sc,   'Main\nScore',          bold=True, fg='FFFFFF', bg=v, size=11)
        bc(ws, 4, sc+1, 'Token-F1',             bold=True, fg='FFFFFF', bg=v, size=11)
        bc(ws, 4, sc+2, 'Soft-F1\n(BERTScore)', bold=True, fg='FFFFFF', bg=v, size=11)
    ws.row_dimensions[4].height = 34

    CAT_BG  = {'Semantic':'E8EAF6','Deterministic':'E8F5E9','N/A':'ECEFF1'}
    CAT_FG  = {'Semantic':PALETTE['sem'],'Deterministic':PALETTE['det'],'N/A':PALETTE['na']}
    CAT_HDR = {'Semantic':PALETTE['sem'],'Deterministic':PALETTE['det'],'N/A':PALETTE['na']}

    data_row = 5; prev_tipo = None
    for campo in CAMPO_ORDER:
        tipo_pt = CAMPO_META.get(campo,{}).get('tipo','N/A')
        tipo    = TIPO_EN.get(tipo_pt, tipo_pt)
        met     = METRICA_EN.get(CAMPO_META.get(campo,{}).get('metrica_principal','N/A'),'N/A')
        met     = 'N/A' if met == '—' else met

        if tipo != prev_tipo:
            ws.merge_cells(f'A{data_row}:{get_column_letter(total_cols)}{data_row}')
            c = ws.cell(row=data_row, column=1, value=CAT_LABEL_EN.get(tipo, tipo))
            c.font = Font(bold=True, color='FFFFFF', name='Arial', size=12)
            c.fill = PatternFill('solid', start_color=CAT_HDR.get(tipo, PALETTE['na']))
            c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            c.border = BORDER_THIN
            ws.row_dimensions[data_row].height = 20
            data_row += 1; prev_tipo = tipo

        bg = PALETTE['alt'] if data_row % 2 == 0 else PALETTE['white']
        bc(ws, data_row, 1, campo, bold=True, halign='left', bg=bg, size=12)
        c_cat = bc(ws, data_row, 2, tipo, bg=CAT_BG.get(tipo, bg), size=11)
        c_cat.font = Font(bold=True, color=CAT_FG.get(tipo,'000000'), name='Arial', size=11)
        bc(ws, data_row, 3, met, halign='left', bg=bg, size=11)

        for vi, ver in enumerate(versions):
            sc = 4+vi*3; r = versions_data[ver].get(campo, {})
            if tipo == 'N/A' or not r or r.get('ausente'):
                for ci in range(sc, sc+3):
                    bc(ws, data_row, ci, 'N/A', bg=PALETTE['blank'], fg='AAAAAA', italic=True)
                continue
            p   = principal_score(campo, r)
            tf1 = r.get('token_f1',{}).get('mean')          if tipo_pt=='Semântico'      else \
                  r.get('exact_match_rate')                   if tipo_pt=='Determinístico' else None
            sf1 = r.get('bertscore_proxy',{}).get('f1_mean') if tipo_pt=='Semântico'      else None
            for ci, val in zip(range(sc, sc+3), [p, tf1, sf1]):
                if val is None: bc(ws, data_row, ci, 'N/A', bg=bg, fg='AAAAAA')
                else:           bc(ws, data_row, ci, val, bold=(ci==sc), bg=score_bg_clean(val) or bg)

        ws.row_dimensions[data_row].height = 20; data_row += 1

    data_row += 1
    ws.merge_cells(f'A{data_row}:{get_column_letter(total_cols)}{data_row}')
    leg = ws.cell(row=data_row, column=1,
        value='Legend:  >= 0.70 Highly Similar  |  0.60-0.70 Moderately Similar  |  '
              '0.40-0.60 Slightly Similar  |  < 0.40 Divergent  |  Absent  '
              '| Main Score = ROUGE-L (semantic) | F1-macro (severity) | '
              'Exact Match (other det.) | Set-F1 (references)')
    leg.font = Font(italic=True, color='555555', name='Arial', size=11)
    leg.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[data_row].height = 18

    ws.column_dimensions['A'].width = 28; ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    for vi in range(n_ver):
        for ci in range(3):
            ws.column_dimensions[get_column_letter(4+vi*3+ci)].width = 13
    ws.freeze_panes = 'A5'

    # ── Charts ────────────────────────────────────────────────────────────
    # Group versions by LLM for per-LLM chart tabs
    from collections import OrderedDict as _OD
    llm_groups = _OD()
    for ver in versions:
        _llm = version_llms.get(ver, ver.split('|')[0] if '|' in ver else ver)
        if _llm not in llm_groups:
            llm_groups[_llm] = []
        llm_groups[_llm].append(ver)

    LLM_TAB_COLORS = {
        'deepseek': '1565C0', 'gpt4': '4A148C', 'gpt5': '2E7D32',
        'llama3': 'BF360C',   'llama4': '00695C',
    }

    charts = [
        ('Overview', '1F3864',
         lambda: chart_overview(versions_data, versions, all_campos, version_llms)),
        ('Overview TMMv1', '1F3864',
         lambda: chart_overview_version('TMMv1', versions_data, versions, all_campos, version_llms)),
        ('Overview TMMv2', '1F3864',
         lambda: chart_overview_version('TMMv2', versions_data, versions, all_campos, version_llms)),
        ('Overview TMMv3', '1F3864',
         lambda: chart_overview_version('TMMv3', versions_data, versions, all_campos, version_llms)),
        ('All LLMs', '006064',
         lambda: chart_all_llms(versions_data, versions, sem_campos, version_llms)),
    ]
    for _ln, _lv in llm_groups.items():
        _color = LLM_TAB_COLORS.get(_ln, '37474F')
        charts.append((
            _ln.upper(), _color,
            (lambda ln=_ln, lv=_lv: chart_llm(ln, lv, versions_data, sem_campos))
        ))
    charts.append((
        'Omission & Halluc.', 'B71C1C',
        lambda: chart_omission_hallucination(
            versions,
            {k: version_dfs.get(k, version_dfs.get(k.split('|')[-1])) for k in versions},
            baseline_df, version_llms),
    ))

    global _CURRENT_FIG_NAME
    for title, bg, fn in charts:
        print(f'  Building {title}...')
        _CURRENT_FIG_NAME = re.sub(r'[^a-z0-9]+', '_', title.lower()).strip('_')
        buf, img_w, img_h = fn()
        _CURRENT_FIG_NAME = None
        make_chart_sheet(wb, title, subtitle, bg, buf, img_w, img_h)

    # ── Omission Summary ──────────────────────────────────────────────────
    print('  Building Omission Summary...')
    vdfs_mapped = {k: version_dfs.get(k, version_dfs.get(k.split('|')[-1])) for k in versions}
    _build_omission_matrix(wb, versions, vdfs_mapped, baseline_df, version_llms)

    wb.save(xlsx_path)
    print(f'\nSaved -> {xlsx_path}')


# ═══════════════════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════════════════

def parse_versions_arg(versions_arg):
    """
    Parse list of 'LABEL:path/to/file.csv' or 'LABEL:path:LLM' strings.
    Returns ordered list of (label, path, llm) tuples. llm may be None.
    """
    result = []
    for item in versions_arg:
        if ':' not in item:
            print(f'ERROR: --versions items must be LABEL:file.csv or LABEL:file.csv:LLM  (got: {item})')
            sys.exit(1)
        parts = item.split(':')
        label = parts[0].strip()
        if len(parts) == 2:
            path, llm = parts[1].strip(), None
        elif len(parts) == 3:
            path, llm = parts[1].strip(), (parts[2].strip() or None)
        else:
            path = ':'.join(parts[1:-1]).strip()
            llm  = parts[-1].strip() or None
        if not Path(path).exists():
            print(f'ERROR: File not found: {path}')
            sys.exit(1)
        result.append((label, path, llm))
    # Sort by LLM name first (so all versions of the same LLM are consecutive),
    # then by version number (V1 < V2 < V3) within each LLM group.
    result.sort(key=lambda x: (x[2] or '', int(x[0][1:]) if x[0][1:].isdigit() else 999))
    return result


def main():
    parser = argparse.ArgumentParser(
        description='TMM — compute metrics and generate xlsx in one command',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Single version
  python tools/TMM_metrics_run.py \\
      --baseline dataset/reports/vulnnet_scans_openvas.csv \\
      --versions TMMv3:dataset/extractions/deepseek_v3.csv:deepseek \\
      --xlsx TMM_metrics_v3.xlsx

  # Multiple versions
  python tools/TMM_metrics_run.py \\
      --baseline dataset/reports/vulnnet_scans_openvas.csv \\
      --versions TMMv3:dataset/extractions/deepseek_v3.csv:deepseek TMMv3:dataset/extractions/gpt4_v3.csv:gpt4 \\
      --xlsx TMM_metrics_v3.xlsx

  # Save DB for later reuse
  python tools/TMM_metrics_run.py \\
      --baseline dataset/reports/vulnnet_scans_openvas.csv \\
      --versions TMMv3:dataset/extractions/deepseek_v3.csv:deepseek \\
      --xlsx TMM_metrics_v3.xlsx \\
      --db metrics_db.json

  # All 5 LLMs, V3 (real repo paths: baselines/native + artifacts/v3/openvas_129_dockers)
  python tools/TMM_metrics_run.py \\
      --baseline baselines/native/vulnnet_scans_openvas.csv \\
      --versions TMMv3:artifacts/v3/openvas_129_dockers/deepseek_v3.csv:deepseek \\
                 TMMv3:artifacts/v3/openvas_129_dockers/gpt4_v3.csv:gpt4 \\
                 TMMv3:artifacts/v3/openvas_129_dockers/gpt5_v3.csv:gpt5 \\
                 TMMv3:artifacts/v3/openvas_129_dockers/llama3_v3.csv:llama3 \\
                 TMMv3:artifacts/v3/openvas_129_dockers/llama4_v3.csv:llama4 \\
      --xlsx artifacts/v3/TMM_metrics_v3.xlsx
        """)

    parser.add_argument('--baseline', required=True,
                        help='Path to baseline CSV (e.g. vulnnet_scans_openvas.csv)')
    parser.add_argument('--versions', required=True, nargs='+', metavar='LABEL:FILE[:LLM]',
                        help='One or more version CSVs in format LABEL:file.csv or LABEL:file.csv:LLM '
                             '(e.g. V1:merged_v1.csv:deepseek V2:merged_v2.csv:gpt4o)')
    parser.add_argument('--xlsx', required=True,
                        help='Output xlsx path')
    parser.add_argument('--db', default=None,
                        help='Optional: save metrics DB as JSON for later reuse')
    parser.add_argument('--figdir', default=None,
                        help='Optional: also export every chart as vector PDF+SVG into this directory')
    parser.add_argument('--paper-layout', action='store_true',
                        help='Compact, vertically-optimized chart layouts for paper figures')

    args = parser.parse_args()

    global FIGDIR, PAPER_LAYOUT
    if args.figdir:
        FIGDIR = Path(args.figdir)
        FIGDIR.mkdir(parents=True, exist_ok=True)
    PAPER_LAYOUT = args.paper_layout

    if not Path(args.baseline).exists():
        print(f'ERROR: Baseline file not found: {args.baseline}')
        sys.exit(1)

    version_pairs = parse_versions_arg(args.versions)
    version_llms  = {lbl: llm for lbl, _, llm in version_pairs}

    def _file_hash(path):
        h = hashlib.md5()
        with open(path, 'rb') as f:
            for chunk in iter(lambda: f.read(65536), b''):
                h.update(chunk)
        return h.hexdigest()

    run_ts    = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    prev_db   = {}

    # Load previous DB for comparison (if --db exists from prior run)
    if args.db and Path(args.db).exists():
        try:
            with open(args.db, encoding='utf-8') as f:
                prev_db = json.load(f)
        except Exception:
            prev_db = {}

    # ── Keys: "LLM|label" guarantees uniqueness across LLMs ────────────
    def _key(label, llm): return f'{llm}|{label}' if llm else label

    prev_hashes      = prev_db.get('file_hashes', {})
    prev_versions    = prev_db.get('versions', {})          # keyed by LLM|label
    prev_ver_order   = prev_db.get('version_order', [])
    prev_ver_llms    = prev_db.get('version_llms', {})
    prev_ver_pairs   = prev_db.get('version_pairs', [])     # [[label,path,llm], ...]
    # Invalidate entire cache if compute logic changed
    if prev_db.get('compute_version') != COMPUTE_VERSION:
        print(f'  [Cache] Compute version changed -> forcing full recompute.')
        prev_hashes = {}; prev_versions = {}

    print(f'\n{"="*60}')
    print(f'  TMM Metrics Pipeline')
    print(f'  Run       : {run_ts}')
    print(f'  Baseline  : {args.baseline}')
    print(f'  Output    : {args.xlsx}')
    print(f'{"="*60}')

    # ── File integrity check & decide what to recompute ──────────────
    print(f'\n  [Source Files]')
    file_hashes = dict(prev_hashes)  # carry over previous entries

    bl_hash = _file_hash(args.baseline)
    prev_bl = prev_hashes.get('baseline', {}).get('md5', '')
    bl_changed = prev_bl and prev_bl != bl_hash
    bl_tag  = '  CHANGED [!]' if bl_changed else ('  unchanged' if prev_bl else '  (first run)')
    print(f'    baseline  {Path(args.baseline).name:<40} md5={bl_hash[:10]}...{bl_tag}')
    file_hashes['baseline'] = {'path': args.baseline, 'md5': bl_hash}

    # Expected semantic fields — if any are missing from cached result, recompute
    _EXPECTED_SEM = {'description','detection_result','detection_method',
                     'product_detection_result','impact','solution','insight','references'}

    to_compute = []   # (label, path, llm, key) — entries that need fresh computation
    for label, path, llm in version_pairs:
        key     = _key(label, llm)
        fh      = _file_hash(path)
        prev_fh = prev_hashes.get(key, {}).get('md5', '')
        cached  = prev_versions.get(key, {})
        incomplete = bool(cached) and not _EXPECTED_SEM.issubset(cached.keys())
        if bl_changed or not prev_fh or prev_fh != fh or incomplete:
            tag = 'CHANGED [!]' if (prev_fh and prev_fh != fh) else \
                  ('INCOMPLETE CACHE — recomputing' if incomplete else
                   ('NEW' if not prev_fh else 'BASELINE CHANGED'))
            to_compute.append((label, path, llm, key))
        else:
            tag = 'unchanged (reusing cached metrics)'
        llm_tag = f'[{llm}]' if llm else ''
        print(f'    {key:<20} {Path(path).name:<36} md5={fh[:10]}... {tag}')
        file_hashes[key] = {'path': path, 'md5': fh, 'llm': llm or '', 'label': label}

    # ── Merge: start from previous accumulated data ───────────────────
    versions_data = dict(prev_versions)   # all previously computed entries
    version_llms_all = dict(prev_ver_llms)
    all_pairs_map = {_key(l, m): (l, p, m) for l, p, m in
                     [tuple(x) for x in prev_ver_pairs]}

    # ── Compute only what is new or changed ───────────────────────────
    if to_compute:
        print(f'\n  [Computing {len(to_compute)} entry/entries]')
        for label, path, llm, key in to_compute:
            print(f'  -> {key}', end='', flush=True)
            results = compute_version(path, args.baseline, label)
            versions_data[key]     = results
            version_llms_all[key]  = llm or ''
            all_pairs_map[key]     = (label, path, llm)
            print(f'  OK')
    else:
        print(f'\n  All entries unchanged — reusing cached metrics.')

    # ── Add current run entries (may already exist, idempotent) ───────
    for label, path, llm in version_pairs:
        key = _key(label, llm)
        version_llms_all[key] = llm or ''
        all_pairs_map[key]    = (label, path, llm)

    # ── Build ordered version list: previous + new, deduped ──────────
    seen = set()
    versions = []
    # First: preserved order from previous run
    for k in prev_ver_order:
        if k in versions_data and k not in seen:
            versions.append(k); seen.add(k)
    # Then: current run entries (new LLMs/versions appended)
    for label, path, llm in version_pairs:
        k = _key(label, llm)
        if k in versions_data and k not in seen:
            versions.append(k); seen.add(k)

    # Build version_pairs_all for build_xlsx (label, path, llm) per key
    version_pairs_all = [all_pairs_map[k] for k in versions if k in all_pairs_map]
    # version_llms for display: key -> llm
    version_llms_disp = {k: all_pairs_map[k][2] or '' for k in versions if k in all_pairs_map}

    print(f'\n  Accumulated dataset: {len(versions)} entry/entries')
    for k in versions:
        llm_d = version_llms_disp.get(k, '')
        print(f'    {k}' + (f' [{llm_d}]' if llm_d else ''))

    # ── Remove stale xlsx before writing ─────────────────────────────
    xlsx_path = Path(args.xlsx)
    if xlsx_path.exists():
        try:
            os.remove(xlsx_path)
            print(f'\n  Previous {xlsx_path.name} removed.')
        except OSError as e:
            print(f'\n  WARNING: could not remove old xlsx: {e}')
            print(f'  Close the file in Excel and retry.')
            sys.exit(1)

    # ── Save accumulated DB ───────────────────────────────────────────
    db_out = {
        'baseline':        args.baseline,
        'compute_version': COMPUTE_VERSION,
        'versions':        versions_data,
        'version_order':   versions,
        'version_llms':    version_llms_all,
        'version_pairs':   [list(all_pairs_map[k]) for k in versions if k in all_pairs_map],
        'campo_order':     CAMPO_ORDER,
        'campo_meta':      CAMPO_META,
        'run_timestamp':   run_ts,
        'file_hashes':     file_hashes,
    }
    db_path = args.db or str(xlsx_path.with_suffix('.db.json'))
    with open(db_path, 'w', encoding='utf-8') as f:
        json.dump(db_out, f, indent=2, ensure_ascii=False)
    print(f'  DB saved -> {db_path}')

    # ── Build xlsx with all accumulated data ─────────────────────────
    print(f'\nBuilding xlsx...')
    build_xlsx(versions_data, versions, args.baseline, args.xlsx,
               version_pairs_all, version_llms_disp)
    print(f'Done.  ->  {args.xlsx}')

    # ── Generate markdown report ──────────────────────────────────────
    md_path = str(xlsx_path.with_suffix('.md'))
    generate_md_report(md_path, versions, versions_data, version_llms_disp,
                       args.baseline, run_ts)
    print(f'Report -> {md_path}')


def generate_md_report(md_path, versions, versions_data, version_llms, baseline, run_ts):
    from collections import defaultdict
    lines = []

    lines += [
        '# TMM — Metrics Report',
        '',
        f'**Generated:** {run_ts}  ',
        f'**Baseline:** `{baseline}`  ',
        f'**Versions:** {len(versions)}',
        '',
        '---',
        '',
        '## 1. Methodology Overview',
        '',
        'Each LLM-generated CSV is matched against the OpenVAS baseline via an **inner join**',
        'on `Name` (LLM CSV) = `NVT Name` (baseline). All rows from both files are used — no',
        'deduplication — so that every scan instance is individually evaluated against every',
        'matching baseline entry. Only matched pairs are scored.',
        '',
        '---',
        '',
        '## 2. Semantic Fields',
        '',
        'Fields whose content is free text and evaluated with similarity metrics:',
        '',
        '| Field | Baseline Column |',
        '|---|---|',
        '| `description` | Summary |',
        '| `detection_result` | Specific Result |',
        '| `detection_method` | Vulnerability Detection Method |',
        '| `product_detection_result` | Product Detection Result |',
        '| `impact` | Impact |',
        '| `solution` | Solution |',
        '| `insight` | Vulnerability Insight |',
        '',
        '### 2.1 Metrics Applied to Each Semantic Field',
        '',
        '| Metric | Description |',
        '|---|---|',
        '| **ROUGE-L** | Longest Common Subsequence F-measure between hypothesis and reference. Range [0, 1]. |',
        '| **Token-F1** | Precision × Recall over token sets (bag-of-words). Range [0, 1]. |',
        '| **TF-IDF Cosine** | Cosine similarity between TF-IDF vectors of hypothesis and reference. Range [0, 1]. |',
        '| **Soft-F1 (BERTScore proxy)** | Word-level TF-IDF soft alignment: for each token in hypothesis find closest token in reference via cosine similarity. Computes Precision (P), Recall (R), and F1. Approximates BERTScore without requiring a neural model. Range [0, 1]. |',
        '',
        '**Main Score** reported in Summary: ROUGE-L mean.',
        '',
        '**Quality thresholds:** >= 0.70 Highly Similar | 0.60–0.70 Moderately Similar | 0.40–0.60 Slightly Similar | < 0.40 Divergent | Absent',
        '',
        '---',
        '',
        '## 3. References Field (CVE IDs)',
        '',
        '`references` maps to the `CVEs` column in the baseline.',
        '',
        '| Metric | Description |',
        '|---|---|',
        '| **Set-F1** | Precision and Recall computed over sets of CVE identifiers extracted from each text. F1 = harmonic mean. |',
        '',
        '---',
        '',
        '## 4. Deterministic Fields',
        '',
        'Fields with a fixed controlled vocabulary or numeric value:',
        '',
        '| Field | Baseline Column | Metric |',
        '|---|---|---|',
        '| `cvss` | CVSS | Exact Match |',
        '| `port` | Port | Exact Match (float→int normalised) |',
        '| `protocol` | Port Protocol | Exact Match |',
        '| `severity` | Severity | F1-macro (multi-class) |',
        '',
        '**Exact Match:** proportion of pairs where the predicted value equals the reference exactly.',
        '',
        '**F1-macro for Severity:** each severity level (LOG / LOW / MEDIUM / HIGH / CRITICAL) is',
        'treated as a class. Precision, Recall, and F1 are computed per class then macro-averaged.',
        '',
        '---',
        '',
        '## 5. Omission & Hallucination',
        '',
        '| Symbol | Condition | Meaning |',
        '|---|---|---|',
        '| **OK** | Both filled | LLM and baseline both have a value |',
        '| **O** (Omission) | LLM empty, baseline filled | LLM failed to produce content for a field that exists in the baseline |',
        '| **A** (Hallucination) | LLM filled, baseline empty | LLM invented content for a field that is empty in the baseline |',
        '| **N/A** | Both empty | Field not applicable for this vulnerability |',
        '',
        'Rate = count / total matched pairs × 100 %.',
        '',
        '---',
        '',
        '## 6. Field Mapping: LLM vs Baseline',
        '',
        'The table below shows how each field in the LLM CSV is mapped to the corresponding',
        'column in the OpenVAS baseline CSV, and which fields exist only in one source.',
        '',
        '### 6.1 Mapped Fields (LLM ↔ Baseline)',
        '',
        '| LLM Field | Baseline Column | Metric Type |',
        '|---|---|---|',
        '| `description` | Summary | Semantic (ROUGE-L, Token-F1, TF-IDF, Soft-F1) |',
        '| `detection_result` | Specific Result | Semantic |',
        '| `detection_method` | Vulnerability Detection Method | Semantic |',
        '| `product_detection_result` | Product Detection Result | Semantic |',
        '| `impact` | Impact | Semantic |',
        '| `solution` | Solution | Semantic |',
        '| `insight` | Vulnerability Insight | Semantic |',
        '| `references` | CVEs | Set-F1 (CVE identifiers) |',
        '| `cvss` | CVSS | Exact Match |',
        '| `port` | Port | Exact Match (numeric) |',
        '| `protocol` | Port Protocol | Exact Match |',
        '| `severity` | Severity | F1-macro (multi-class) |',
        '',
        '### 6.2 LLM-Only Fields (no baseline counterpart)',
        '',
        '| LLM Field | Description |',
        '|---|---|',
        '| `log_method` | Logging method used by the LLM during extraction |',
        '| `plugin` | Plugin identifier used in extraction |',
        '| `identification` | Identification details extracted by LLM |',
        '| `http_info` | HTTP-specific information extracted by LLM |',
        '| `source` | Source of the vulnerability data |',
        '| `llm` | LLM model identifier used for extraction |',
        '| `target` | Target host information |',
        '',
        '### 6.3 Baseline-Only Columns (not in LLM CSV)',
        '',
        '| Baseline Column | Description |',
        '|---|---|',
        '| IP | Host IP address |',
        '| Hostname | Host FQDN or alias |',
        '| QoD | Quality of Detection score |',
        '| Solution Type | Type of solution (Mitigation / VendorFix / etc.) |',
        '| NVT OID | OpenVAS NVT Object Identifier |',
        '| Task ID | Scan task identifier |',
        '| Task Name | Scan task name |',
        '| Timestamp | Scan result timestamp |',
        '| Result ID | Unique result identifier |',
        '| Affected Software/OS | Affected software or OS version |',
        '| BIDs | Bugtraq IDs |',
        '| CERTs | CERT advisories |',
        '| Other References | Additional vulnerability references |',
        '| NVT Name | Vulnerability name (join key) |',
        '',
        '---',
        '',
        '## 7. Results Summary',
        '',
    ]

    # Group by LLM
    llm_groups = defaultdict(list)
    for ver in versions:
        llm = version_llms.get(ver, ver.split('|')[0] if '|' in ver else ver)
        llm_groups[llm].append(ver)

    sem_campos = [c for c in CAMPO_ORDER
                  if CAMPO_META.get(c, {}).get('tipo') == 'Semântico' and c != 'references']
    det_campos = ['cvss', 'port', 'protocol', 'severity', 'references']

    for llm, vers in llm_groups.items():
        lines += [f'### {llm.upper()}', '']
        lines += ['**Semantic — ROUGE-L mean**', '']
        hdr = '| Version | ' + ' | '.join(sem_campos) + ' |'
        sep = '|---|' + '---|' * len(sem_campos)
        lines += [hdr, sep]
        for ver in vers:
            vd  = versions_data.get(ver, {})
            lbl = ver.split('|')[-1] if '|' in ver else ver
            vals = [vd.get(c, {}).get('rouge_l', {}).get('mean') for c in sem_campos]
            row  = f'| {lbl} | ' + ' | '.join(f'{v:.4f}' if v is not None else 'N/A' for v in vals) + ' |'
            lines.append(row)
        lines.append('')
        lines += ['**Deterministic scores**', '']
        hdr2 = '| Version | CVSS EM | Port EM | Protocol EM | Severity F1 | CVE Set-F1 |'
        lines += [hdr2, '|---|---|---|---|---|---|']
        for ver in vers:
            vd  = versions_data.get(ver, {})
            lbl = ver.split('|')[-1] if '|' in ver else ver
            cvss = vd.get('cvss', {}).get('exact_match_rate')
            port = vd.get('port', {}).get('exact_match_rate')
            prot = vd.get('protocol', {}).get('exact_match_rate')
            sev  = vd.get('severity', {}).get('f1_macro')
            cve  = vd.get('references', {}).get('set_f1_cve', {}).get('mean')
            fmt  = lambda v: f'{v:.4f}' if v is not None else 'N/A'
            lines.append(f'| {lbl} | {fmt(cvss)} | {fmt(port)} | {fmt(prot)} | {fmt(sev)} | {fmt(cve)} |')
        lines += ['', '---', '']

    with open(md_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))


if __name__ == '__main__':
    main()